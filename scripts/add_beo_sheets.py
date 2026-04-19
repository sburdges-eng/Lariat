#!/usr/bin/env python3
"""Add BEO Menu, BEO Event, and BEO Prep sheets to the Lariat Operations Workbook."""
import sys
from pathlib import Path
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

TODAY = date.today().isoformat()
WB_PATH = ROOT / f"lariat_operations_workbook_{TODAY}.xlsx"

# ── Style constants ──
DARK_GREEN = "1B4332"
MED_GREEN = "2D6A4F"
LIGHT_GREEN = "D8F3DC"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY = "D9D9D9"
CREAM = "FFF8F0"
GOLD = "F4A261"
NAVY = "264653"
RUST = "9B2226"
WARM_BG = "FEF3E2"
BLUE_INPUT = "0000FF"

MONEY_FMT = '$#,##0.00'
THIN_BORDER = Border(
    bottom=Side(style="thin", color=MED_GRAY),
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
)
ALL_THIN = Border(
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
)


# ═══════════════════════════════════════════════════════════════
#  BEO MENU CATALOG — master reference for all catering items
# ═══════════════════════════════════════════════════════════════

BEO_MENU = [
    # (item, price, category, recipe_link, prep_notes, pre_prep_notes, plating_notes)
    # ── Passed Apps ──
    ("Mini Rellenos", 4, "Passed App", "", "Stuff & Bread", "Roast peppers, make filling", "4 inch plastic"),
    ("Deviled Eggs", 4, "Passed App", "", "Pipe filling", "Hard boil eggs, make filling", "4 inch plastic"),
    ("Caprese Skewers", 4, "Passed App", "", "Assemble (Cherry tom-Basil-Ciliegine, Prosciutto)", "NONE", "Platter"),
    ("Mac Balls", 4, "Passed App", "", "Roll into 4oz ball, Par-Cook", "Par-Cook Ditalini, Mise Panko Egg Wash Flour", "4 inch plastic"),
    ("Corn Dogs", 4, "Passed App", "corndog_batter", "Batter & fry", "Make corndog batter", "Baskets"),
    ("Pig Wings", 4, "Passed App", "alabama_white_sauce", "Fry & sauce", "Brine pig wings, make Alabama white sauce", "4 inch plastic"),
    ("Bri and Raspberry Philo Bites", 4, "Passed App", "", "NONE", "NONE", "4 inch plastic"),
    ("Cowboy Corn Bites", 4, "Passed App", "", "Assemble & grill", "Roast corn, make filling", "4 inch plastic"),
    ("Texmex Eggrolls", 4, "Passed App", "", "Roll & fry", "Make filling, prep wrappers", "4 inch plastic"),
    ("Caprese Toast", 6, "Passed App", "", "Assemble & toast", "Slice bread, prep toppings", "Platter"),
    # ── Soups ──
    ("Gazpacho", 5, "Soup", "", "Blend & chill", "Dice vegetables, chill overnight", "Cups"),
    ("Chilled Corn Leek", 5, "Soup", "", "Cook & chill", "Sweat leeks, cook corn", "Cups"),
    ("Watermelon Gazpacho", 6, "Soup", "", "Blend & chill", "Cube watermelon, prep garnish", "Cups"),
    # ── Tacos & Sliders ──
    ("Braised Chicken Taco", 6, "Taco/Slider", "", "Build tacos", "Braise chicken, make slaw, warm tortillas", "NONE"),
    ("Pork Belly Bao Bun", 6, "Taco/Slider", "", "Steam buns & assemble", "Braise pork belly, pickle veg", "NONE"),
    ("Carnitas Taco", 6, "Taco/Slider", "", "Build tacos", "Braise carnitas, make pico, slaw", "NONE"),
    ("Battered Avocado Taco", 6, "Taco/Slider", "beer_batter", "Batter & fry avocado, build", "Make beer batter, slice avocado, prep slaw", "NONE"),
    ("Nashville Slider", 6, "Taco/Slider", "nashville_hot_rub", "Fry & sauce, build slider", "Brine chicken, make Nashville rub & oil, pickle slaw", "NONE"),
    ("Rope Burger Slider", 7, "Taco/Slider", "bacon_jam", "Grill & build slider", "Portion patties, make bacon jam, prep toppings", "NONE"),
    ("Barbacoa Taco", 7, "Taco/Slider", "birria", "Build tacos", "Braise beef cheeks 8-12hr, strain consomme", "NONE"),
    ("Battered Fish Taco", 7, "Taco/Slider", "beer_batter", "Prepare Batter + Flour Dredge, fry & build", "Brine & Cut Fish, Pico, Red Slaw", "NONE"),
    ("Prime Rib Sliders", 10, "Taco/Slider", "", "Slice & build slider", "Season & cook prime rib, horseradish cream", "NONE"),
    # ── Premium Apps ──
    ("Crab Cake Remoulade", 12, "Premium App", "", "Form & sear cakes", "Make crab cake mix, remoulade", "Platter"),
    ("Beef Tenderloin Crostini", 11, "Premium App", "", "Sear, slice, build", "Trim & tie tenderloin, make crostini, horseradish", "Platter"),
    ("Trio Dips", 30, "Shared Platter", "queso_mac_sauce", "Green Chile, Black Salsa, Queso, Chips", "Make green chile, blackened salsa, queso", "Usual — 3 bowls + chips"),
    # ── Buffets ──
    ("Rope Caesar Salad Buffet", 150, "Buffet", "santa_fe_caesar", "Assemble salad buffet", "Cook Cornbread, Caesar Dressing, Grill Red Onions, Make Succotash", "Full pan"),
    ("Cobb Salad Buffet", 150, "Buffet", "cobb_dressing", "Assemble salad buffet", "Make cobb dressing, grill chicken, cook eggs & bacon", "Full pan"),
    ("Green Chile Mac Buffet", 125, "Buffet", "queso_mac_sauce", "Cook mac, hold hot", "Make queso sauce, par-cook pasta, make green chile", "Full pan"),
    ("Carnitas Tacos Buffet", 250, "Buffet", "", "Build taco station", "Braise carnitas, make all salsas, slaw, warm tortillas", "Full pan station"),
    ("Battered Avocado Taco Buffet", 250, "Buffet", "beer_batter", "Fry station", "Make beer batter, slice avocado, prep slaw & pico", "Full pan station"),
    ("Braised Chicken Taco Buffet", 250, "Buffet", "", "Build taco station", "Braise chicken, make slaw & salsas, warm tortillas", "Full pan station"),
    ("Barbacoa Taco Buffet", 300, "Buffet", "birria", "Build taco station", "Braise beef cheeks 8-12hr, strain consomme, salsas", "Full pan station"),
    ("Fish Taco Buffet", 300, "Buffet", "beer_batter", "Fry station", "Brine & Cut Fish, Pico, Red Slaw, Beer Batter", "Full pan station"),
    ("Low Country Boil", 20, "Buffet", "", "Boil & serve", "Prep shrimp, sausage, corn, potatoes, season water", "Station"),
    # ── Artisanal Boards ──
    ("French Artisanal Board", 200, "Board", "", "Assemble board", "Source cheeses, charcuterie, accoutrements", "Board"),
    ("Italian Artisanal Board", 200, "Board", "", "Assemble board", "Source cheeses, charcuterie, accoutrements", "Board"),
    ("Spanish Artisanal Board", 200, "Board", "", "Assemble board", "Source cheeses, charcuterie, accoutrements", "Board"),
    # ── Desserts ──
    ("Churros", 4, "Dessert", "", "Fry & coat", "Make churro dough, prep dulce/chocolate", "Baskets"),
    ("Cupcakes", 5, "Dessert", "", "NONE", "NONE", "Platter"),
    ("Chocolate Cake", 7, "Dessert", "", "Slice & plate", "Bake cake, make ganache", "Platter"),
    ("Banana Cream Pudding", 5, "Dessert", "", "Assemble & chill", "Make pudding, slice bananas, prep wafers", "Cups"),
    ("Tiramisu", 7, "Dessert", "", "Assemble & chill", "Make mascarpone filling, brew espresso, dip ladyfingers", "Full pan"),
    # ── Dinner Packages ──
    ("Roast Chicken Dinner", 80, "Dinner Package", "chicken_confit", "Full dinner service", "Full prep per package menu", "Plated"),
    ("Mexican Dinner", 80, "Dinner Package", "", "Full dinner service", "Full prep per package menu", "Family style"),
    ("Italian Dinner", 80, "Dinner Package", "", "Full dinner service", "Full prep per package menu", "Family style"),
    ("Prime Rib Dinner", 95, "Dinner Package", "", "Full dinner service", "Season & cook prime rib, sides per package", "Plated"),
    # ── Entrees / Stations (per-unit pricing) ──
    ("Roasted Chicken Leg", 225, "Entree Station", "chicken_confit", "Confit & roast", "French legs, confit EVOO 225F 3hr", "Full pan"),
    ("Roasted Root Veg", 175, "Entree Station", "", "Roast", "Cut root veg, toss with oil & herbs", "Full pan"),
    ("Baked Ziti", 175, "Entree Station", "", "Cook", "Par-Cook Ziti, Marinara", "Full pan"),
    ("Prime Rib Carving Station", 35, "Entree Station", "", "Sous-Vide @ 9:00, sear & carve", "Season & Seal Prime Rib", "Station"),
    # ── Bar ──
    ("OPEN BAR BASIC", 20, "Bar", "", "", "", ""),
    ("Bar Spend Amount", 0, "Bar", "", "", "", ""),
]


def build_beo_menu_sheet(wb):
    if "BEO Menu" in wb.sheetnames:
        del wb["BEO Menu"]
    ws = wb.create_sheet(title="BEO Menu")

    ws.cell(row=1, column=1, value="THE LARIAT — BEO CATERING MENU")
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=16, color=DARK_GREEN)
    ws.merge_cells("A1:G1")

    headers = ["Item", "Price", "Category", "Recipe Link", "Prep", "Pre-Prep", "Plating"]
    hr = 3
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A4"

    r = hr + 1
    current_cat = ""
    for item, price, cat, recipe, prep, preprep, plating in BEO_MENU:
        if cat != current_cat:
            ws.cell(row=r, column=1, value=cat.upper())
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            for c in range(1, 8):
                ws.cell(row=r, column=c).font = Font(name="Arial", bold=True, size=10, color=WHITE)
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GOLD)
            current_cat = cat
            r += 1

        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=price)
        ws.cell(row=r, column=2).number_format = MONEY_FMT
        ws.cell(row=r, column=3, value=cat)
        if recipe:
            ws.cell(row=r, column=4, value=recipe)
            safe = recipe[:31]
            if safe in wb.sheetnames:
                ws.cell(row=r, column=4).hyperlink = f"#'{safe}'!A1"
                ws.cell(row=r, column=4).font = Font(name="Arial", size=10, color="0563C1", underline="single")
        ws.cell(row=r, column=5, value=prep)
        ws.cell(row=r, column=6, value=preprep)
        ws.cell(row=r, column=7, value=plating)

        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            if not cell.font.bold and not cell.font.underline:
                cell.font = Font(name="Arial", size=10)
            cell.border = ALL_THIN
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        if (r - hr) % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=WARM_BG)
        r += 1

    widths = [34, 10, 16, 18, 36, 40, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.print_title_rows = "1:3"
    last_data_row = r - 1
    print(f"  BEO Menu: {last_data_row - hr} items across {len(set(m[2] for m in BEO_MENU))} categories")
    return hr, last_data_row


def build_beo_event_sheet(wb, menu_hr, menu_last_row):
    if "BEO Event" in wb.sheetnames:
        del wb["BEO Event"]
    ws = wb.create_sheet(title="BEO Event")

    ws.cell(row=1, column=1, value="BEO EVENT CALCULATOR")
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=16, color=DARK_GREEN)
    ws.merge_cells("A1:H1")

    # ── Event info inputs ──
    info = [
        ("Event Name:", "B2", ""),
        ("Client:", "B3", ""),
        ("Date:", "D2", ""),
        ("Guest Count:", "D3", 20),
        ("Event ID:", "F2", ""),
        ("Venue:", "F3", ""),
    ]
    labels = [("A2", "Event Name:"), ("A3", "Client:"), ("C2", "Date:"), ("C3", "Guest Count:"),
              ("E2", "Event ID:"), ("E3", "Venue:")]
    for cell_ref, label in labels:
        ws[cell_ref] = label
        ws[cell_ref].font = Font(name="Arial", bold=True, size=10)

    ws["B2"].font = Font(name="Arial", size=10, color=BLUE_INPUT)
    ws["B3"].font = Font(name="Arial", size=10, color=BLUE_INPUT)
    ws["D2"].font = Font(name="Arial", size=10, color=BLUE_INPUT)
    ws["D3"] = 20
    ws["D3"].font = Font(name="Arial", bold=True, size=12, color=BLUE_INPUT)
    ws["D3"].fill = PatternFill("solid", fgColor="FFFF00")
    ws["F2"].font = Font(name="Arial", size=10, color=BLUE_INPUT)
    ws["F3"].font = Font(name="Arial", size=10, color=BLUE_INPUT)

    ws.cell(row=4, column=1)

    # ── Column headers ──
    headers = ["Item", "Unit Price", "Qty", "Line Total", "Category", "Prep", "Pre-Prep", "Plating"]
    hr = 5
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A6"

    # ── Item entry rows (40 rows for event items) ──
    MAX_ITEMS = 40
    menu_range = f"'BEO Menu'!$A${menu_hr+1}:$A${menu_last_row}"
    item_dv = DataValidation(type="list", formula1=f"={menu_range}", allow_blank=True)
    item_dv.errorTitle = "Invalid Item"
    item_dv.error = "Select from BEO Menu"
    ws.add_data_validation(item_dv)

    for r in range(hr + 1, hr + 1 + MAX_ITEMS):
        item_cell = ws.cell(row=r, column=1)
        item_dv.add(item_cell)
        item_cell.font = Font(name="Arial", size=10, color=BLUE_INPUT)

        # B: Unit Price = VLOOKUP to BEO Menu
        a_ref = f"A{r}"
        menu_table = f"'BEO Menu'!$A${menu_hr+1}:$G${menu_last_row}"
        ws.cell(row=r, column=2).value = f'=IF({a_ref}="","",VLOOKUP({a_ref},{menu_table},2,FALSE))'
        ws.cell(row=r, column=2).number_format = MONEY_FMT

        # C: Qty (manual input)
        ws.cell(row=r, column=3).font = Font(name="Arial", size=10, color=BLUE_INPUT)

        # D: Line Total = Price × Qty
        ws.cell(row=r, column=4).value = f'=IF(OR(B{r}="",C{r}=""),"",B{r}*C{r})'
        ws.cell(row=r, column=4).number_format = MONEY_FMT

        # E: Category
        ws.cell(row=r, column=5).value = f'=IF({a_ref}="","",VLOOKUP({a_ref},{menu_table},3,FALSE))'

        # F: Prep
        ws.cell(row=r, column=6).value = f'=IF({a_ref}="","",VLOOKUP({a_ref},{menu_table},5,FALSE))'
        ws.cell(row=r, column=6).alignment = Alignment(wrap_text=True)

        # G: Pre-Prep
        ws.cell(row=r, column=7).value = f'=IF({a_ref}="","",VLOOKUP({a_ref},{menu_table},6,FALSE))'
        ws.cell(row=r, column=7).alignment = Alignment(wrap_text=True)

        # H: Plating
        ws.cell(row=r, column=8).value = f'=IF({a_ref}="","",VLOOKUP({a_ref},{menu_table},7,FALSE))'

        for c in range(1, 9):
            ws.cell(row=r, column=c).border = ALL_THIN

    # ── Pre-fill with the sample event data ──
    sample_event = [
        ("Mac Balls", 20),
        ("Battered Fish Taco", 20),
        ("Caprese Skewers", 20),
        ("Bri and Raspberry Philo Bites", 20),
        ("Rope Caesar Salad Buffet", 2),
        ("Fish Taco Buffet", 1),
        ("Baked Ziti", 1),
        ("Prime Rib Carving Station", 25),
        ("Churros", 25),
        ("Cupcakes", 24),
        ("Trio Dips", 8),
    ]
    for i, (item, qty) in enumerate(sample_event):
        r = hr + 1 + i
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=3, value=qty)

    # ── Totals section ──
    totals_row = hr + 1 + MAX_ITEMS + 1
    d_col = get_column_letter(4)
    first = hr + 1
    last = hr + MAX_ITEMS

    ws.cell(row=totals_row, column=3, value="Subtotal:")
    ws.cell(row=totals_row, column=3).font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=totals_row, column=4).value = f"=SUM({d_col}{first}:{d_col}{last})"
    ws.cell(row=totals_row, column=4).number_format = MONEY_FMT
    ws.cell(row=totals_row, column=4).font = Font(name="Arial", bold=True, size=11)

    ws.cell(row=totals_row + 1, column=3, value="Tax (8.16%):")
    ws.cell(row=totals_row + 1, column=3).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=totals_row + 1, column=4).value = f"={d_col}{totals_row}*0.0816"
    ws.cell(row=totals_row + 1, column=4).number_format = MONEY_FMT

    ws.cell(row=totals_row + 2, column=3, value="Service Fee (20%):")
    ws.cell(row=totals_row + 2, column=3).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=totals_row + 2, column=4).value = f"={d_col}{totals_row}*0.20"
    ws.cell(row=totals_row + 2, column=4).number_format = MONEY_FMT

    ws.cell(row=totals_row + 3, column=3, value="TOTAL:")
    ws.cell(row=totals_row + 3, column=3).font = Font(name="Arial", bold=True, size=12, color=DARK_GREEN)
    ws.cell(row=totals_row + 3, column=4).value = f"={d_col}{totals_row}+{d_col}{totals_row+1}+{d_col}{totals_row+2}"
    ws.cell(row=totals_row + 3, column=4).number_format = MONEY_FMT
    ws.cell(row=totals_row + 3, column=4).font = Font(name="Arial", bold=True, size=12, color=DARK_GREEN)
    for c in range(3, 5):
        ws.cell(row=totals_row + 3, column=c).border = Border(top=Side(style="double", color=DARK_GREEN), bottom=Side(style="double", color=DARK_GREEN))

    # ── Per guest ──
    ws.cell(row=totals_row + 5, column=3, value="Cost / Guest:")
    ws.cell(row=totals_row + 5, column=3).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=totals_row + 5, column=4).value = f'=IF($D$3>0,{d_col}{totals_row+3}/$D$3,"")'
    ws.cell(row=totals_row + 5, column=4).number_format = MONEY_FMT
    ws.cell(row=totals_row + 5, column=4).font = Font(name="Arial", bold=True, size=10)

    widths = [34, 12, 8, 12, 16, 36, 40, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.print_title_rows = "1:5"
    print(f"  BEO Event: {MAX_ITEMS} item slots, {len(sample_event)} pre-filled, totals at row {totals_row}")
    return hr, MAX_ITEMS


def build_beo_prep_sheet(wb, event_hr, max_items):
    if "BEO Prep" in wb.sheetnames:
        del wb["BEO Prep"]
    ws = wb.create_sheet(title="BEO Prep")

    ws.cell(row=1, column=1, value="BEO KITCHEN PREP SHEET")
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=16, color=DARK_GREEN)
    ws.merge_cells("A1:I1")

    ws.cell(row=2, column=1, value="Auto-populated from BEO Event sheet — add items there, prep list updates here")
    ws.cell(row=2, column=1).font = Font(name="Arial", italic=True, size=9, color="888888")
    ws.merge_cells("A2:I2")

    # ── Scale multiplier ──
    ws.cell(row=3, column=1, value="Prep Scale →")
    ws.cell(row=3, column=1).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=3, column=2, value=1.0)
    ws.cell(row=3, column=2).font = Font(name="Arial", bold=True, size=12, color=BLUE_INPUT)
    ws.cell(row=3, column=2).fill = PatternFill("solid", fgColor="FFFF00")
    ws.cell(row=3, column=2).number_format = '0.0'

    dv = DataValidation(type="decimal", operator="greaterThan", formula1="0")
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=3, column=2))

    ws.cell(row=3, column=4, value="Guest Count:")
    ws.cell(row=3, column=4).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=3, column=5).value = "='BEO Event'!D3"
    ws.cell(row=3, column=5).font = Font(name="Arial", bold=True, size=12, color=DARK_GREEN)

    # ── Headers ──
    headers = ["✓", "Item", "Qty", "Scaled Qty", "Category", "PREP", "PRE-PREP", "PLATING", "NOTES"]
    hr = 5
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=RUST)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hr].height = 28
    ws.freeze_panes = "A6"

    # ── Checkbox ──
    cb_dv = DataValidation(type="list", formula1='"✓"', allow_blank=True)
    ws.add_data_validation(cb_dv)

    SCALE_CELL = "$B$3"

    for i in range(max_items):
        r = hr + 1 + i
        event_row = event_hr + 1 + i
        item_ref = f"'BEO Event'!A{event_row}"
        qty_ref = f"'BEO Event'!C{event_row}"

        # A: Checkbox
        ws.cell(row=r, column=1, value="")
        cb_dv.add(ws.cell(row=r, column=1))
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")

        # B: Item (linked from event)
        ws.cell(row=r, column=2).value = f'=IF({item_ref}="","",{item_ref})'

        # C: Qty
        ws.cell(row=r, column=3).value = f'=IF({qty_ref}="","",{qty_ref})'

        # D: Scaled Qty
        ws.cell(row=r, column=4).value = f'=IF(C{r}="","",C{r}*{SCALE_CELL})'
        ws.cell(row=r, column=4).font = Font(name="Arial", bold=True, size=10)

        # E: Category
        ws.cell(row=r, column=5).value = f"='BEO Event'!E{event_row}"

        # F: Prep
        ws.cell(row=r, column=6).value = f"='BEO Event'!F{event_row}"
        ws.cell(row=r, column=6).alignment = Alignment(wrap_text=True)

        # G: Pre-Prep
        ws.cell(row=r, column=7).value = f"='BEO Event'!G{event_row}"
        ws.cell(row=r, column=7).alignment = Alignment(wrap_text=True)

        # H: Plating
        ws.cell(row=r, column=8).value = f"='BEO Event'!H{event_row}"

        # I: Notes (free-text for BOH to add notes during prep)
        ws.cell(row=r, column=9, value="")
        ws.cell(row=r, column=9).font = Font(name="Arial", size=10, color=BLUE_INPUT)

        for c in range(1, 10):
            ws.cell(row=r, column=c).border = ALL_THIN
            if not ws.cell(row=r, column=c).font.bold:
                ws.cell(row=r, column=c).font = Font(name="Arial", size=10)

    # ── Conditional row shading info ──
    ws.cell(row=hr + max_items + 2, column=1, value="Rows auto-hide when no item is selected on BEO Event sheet.")
    ws.cell(row=hr + max_items + 2, column=1).font = Font(name="Arial", italic=True, size=9, color="888888")

    widths = [4, 34, 8, 10, 16, 36, 40, 18, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    print(f"  BEO Prep: {max_items} rows linked to BEO Event, checkbox + scale multiplier")


def main():
    print(f"Loading workbook: {WB_PATH.name}")
    wb = load_workbook(str(WB_PATH))
    print(f"  Existing sheets: {wb.sheetnames}")

    print("Building BEO Menu sheet...")
    menu_hr, menu_last_row = build_beo_menu_sheet(wb)

    print("Building BEO Event sheet...")
    event_hr, max_items = build_beo_event_sheet(wb, menu_hr, menu_last_row)

    print("Building BEO Prep sheet...")
    build_beo_prep_sheet(wb, event_hr, max_items)

    print(f"Saving workbook...")
    wb.save(str(WB_PATH))
    print(f"Done — BEO sheets added to {WB_PATH.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
