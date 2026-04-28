#!/usr/bin/env python3
"""Pure parser for Lariat Shows MKT Plan xlsx → JSON on stdout.

Exit codes:
  0  ok
  2  xlsx_not_found
  3  xlsx_locked (Excel `~$` lock file present)
  4  unrecoverable parse error

Output shape (stdout):
  {
    "shows":          [{band_name, show_date, price, door_tix, status: {...}, source_row}],
    "shows_archive":  [{band_name, show_date, era_year, source_row}],
    "tiktok_ideas":   [{idea, video_content, staff_needed, props, notes, source_row}],
    "dropped":        [{sheet, source_row, reason}]
  }

No DB writes. No side effects beyond stdout.
"""
from __future__ import annotations

import datetime as _dt
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

STATUS_KEYS = [
    "media_list", "mkting_adv", "auto_counts", "announce_date", "meta_ads",
    "fb_event", "co_host_sent", "create_dice_tickets",
    "listing_jambase_bit_songkick", "dice_email", "newsletter",
    "assets", "posts", "whbv",
]


def _emit(payload: dict[str, Any], code: int = 0) -> None:
    print(json.dumps(payload, default=str))
    sys.exit(code)


def _iso(d: Any) -> str | None:
    if isinstance(d, _dt.datetime):
        return d.date().isoformat()
    if isinstance(d, _dt.date):
        return d.isoformat()
    return None


def _str_or_none(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _parse_future(ws) -> tuple[list[dict], list[dict]]:
    shows: list[dict] = []
    dropped: list[dict] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Pad/truncate to length 18 (col 19 is the trailing None we ignore).
        cells = list(row)[:18] + [None] * max(0, 18 - len(row))
        band, date, *rest = cells
        if not band or not isinstance(band, str) or not band.strip():
            dropped.append({"sheet": "future", "source_row": idx, "reason": "missing band_name"})
            continue
        iso = _iso(date)
        if not iso:
            dropped.append({"sheet": "future", "source_row": idx, "reason": "missing or invalid date"})
            continue
        # Columns 3..16 (14 status cells), 17 = door_tix, 18 = price.
        status_cells = rest[:14]
        door_tix = _str_or_none(rest[14]) if len(rest) > 14 else None
        price = rest[15] if len(rest) > 15 else None
        if isinstance(price, str):
            try:
                price = float(price)
            except ValueError:
                price = None
        status = {
            STATUS_KEYS[i]: ("" if v is None else str(v).strip())
            for i, v in enumerate(status_cells)
        }
        shows.append({
            "band_name": band.strip(),
            "show_date": iso,
            "price": price if isinstance(price, (int, float)) else None,
            "door_tix": door_tix,
            "status": status,
            "source_row": idx,
        })
    return shows, dropped


def _parse_past(ws) -> tuple[list[dict], list[dict]]:
    archive: list[dict] = []
    dropped: list[dict] = []
    current_year: int | None = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        a = row[0] if len(row) > 0 else None
        c = row[2] if len(row) > 2 else None

        # Year banner: int/float-ish in col A and no date in col C
        if isinstance(a, (int, float)) and not isinstance(c, _dt.date):
            current_year = int(a)
            continue
        if isinstance(a, str) and re.fullmatch(r"\d{4}", a.strip()) and not isinstance(c, _dt.date):
            current_year = int(a.strip())
            continue

        # Data row: col A is band string, col C is date.
        if isinstance(a, str) and isinstance(c, (_dt.datetime, _dt.date)):
            archive.append({
                "band_name": a.strip(),
                "show_date": _iso(c),
                "era_year": current_year,
                "source_row": idx,
            })
            continue

        # Skip blanks silently; capture genuinely malformed (band but no date).
        if isinstance(a, str) and a.strip():
            dropped.append({
                "sheet": "past", "source_row": idx,
                "reason": "missing or invalid date for band",
            })
    return archive, dropped


def _parse_tiktok(ws) -> tuple[list[dict], list[dict]]:
    out: list[dict] = []
    dropped: list[dict] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = list(row)[:5] + [None] * max(0, 5 - len(row))
        idea, video, staff, props, notes = cells
        idea_s = _str_or_none(idea)
        if not idea_s:
            continue
        out.append({
            "idea": idea_s,
            "video_content": _str_or_none(video),
            "staff_needed": _str_or_none(staff),
            "props": _str_or_none(props),
            "notes": _str_or_none(notes),
            "source_row": idx,
        })
    return out, dropped


def main(argv: list[str]) -> None:
    if len(argv) < 2:
        _emit({"error": "usage", "msg": "ingest_shows_xlsx.py <path-to-xlsx>"}, code=2)
    path = Path(argv[1])
    if not path.exists():
        _emit({"error": "xlsx_not_found", "path": str(path)}, code=2)
    lock = path.parent / f"~${path.name}"
    if lock.exists():
        _emit({"error": "xlsx_locked", "lock": str(lock)}, code=3)

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        future_ws = wb["future"] if "future" in wb.sheetnames else None
        past_ws = wb["past"] if "past" in wb.sheetnames else None
        tiktok_ws = wb["tiktok plan"] if "tiktok plan" in wb.sheetnames else None

        shows, d1 = _parse_future(future_ws) if future_ws else ([], [])
        archive, d2 = _parse_past(past_ws) if past_ws else ([], [])
        tiktok, d3 = _parse_tiktok(tiktok_ws) if tiktok_ws else ([], [])
        wb.close()
    except Exception as e:  # parser failure
        _emit({"error": "parse_error", "msg": str(e)}, code=4)

    _emit({
        "shows": shows,
        "shows_archive": archive,
        "tiktok_ideas": tiktok,
        "dropped": d1 + d2 + d3,
    }, code=0)


if __name__ == "__main__":
    main(sys.argv)
