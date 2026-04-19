#!/usr/bin/env python3
"""Add a Prep Sheet tab to the Lariat Operations Workbook."""
import sys
from pathlib import Path
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

TODAY = date.today().isoformat()
WB_PATH = ROOT / f"lariat_operations_workbook_{TODAY}.xlsx"

# ── Style constants ──
DARK_GREEN = "1B4332"
MED_GREEN = "2D6A4F"
LIGHT_GREEN = "D8F3DC"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY = "D9D9D9"
CREAM = "FFF8F0"
SUB_BG = "EDF2F4"
SUB2_BG = "E2E8F0"
STATION_COLORS = {
    "grill": "E76F51",
    "fry": "F4A261",
    "saute": "E9C46A",
    "salad": "2A9D8F",
    "expo": "264653",
    "all": "6C757D",
}
STATION_LABELS = {
    "grill": "GRILL",
    "fry": "FRY",
    "saute": "SAUTÉ",
    "salad": "SALAD / COLD",
    "expo": "EXPO",
    "all": "ALL STATIONS",
}
MONEY_FMT = '$#,##0.00'
THIN_BORDER = Border(
    bottom=Side(style="thin", color=MED_GRAY),
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
)
THICK_BOTTOM = Border(bottom=Side(style="medium", color=DARK_GREEN))


def safe_text(val):
    if pd.notna(val) and isinstance(val, str) and val.startswith("="):
        return "'" + val
    return val


def load_data():
    recipe_idx = pd.read_csv(ROOT / "recipes" / "recipe_index.csv")
    bom = pd.read_csv(ROOT / "costing" / "bom_2026-04-05.csv")
    recipes = {}
    for f in (ROOT / "recipes" / "normalized").glob("*.csv"):
        recipes[f.stem] = pd.read_csv(f)
    return recipe_idx, bom, recipes


def get_sub_recipes(recipe_idx):
    subs = {}
    for _, row in recipe_idx.iterrows():
        sr = row.get("sub_recipes", "")
        if pd.notna(sr) and sr:
            subs[row["recipe_id"]] = [s.strip() for s in str(sr).split(";") if s.strip()]
        else:
            subs[row["recipe_id"]] = []
    return subs


def build_bom_lookup(bom):
    lookup = {}
    for _, row in bom.iterrows():
        rid = row["recipe_id"]
        if rid not in lookup:
            lookup[rid] = {}
        lookup[rid][row["ingredient"]] = {
            "ext_cost_usd": row.get("ext_cost_usd", None),
        }
    return lookup


def normalize_station(raw):
    if pd.isna(raw) or not raw:
        return ["all"]
    return [s.strip().lower() for s in str(raw).split(";")]


def write_prep_headers(ws):
    headers = [
        "✓", "Recipe / Ingredient", "Category", "Base Qty", "Unit",
        "Multiplier", "Scaled Qty", "Yield", "Yield Unit",
        "Method / Notes", "Batch Cost", "Scaled Cost"
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=DARK_GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = "A5"
    widths = [4, 36, 11, 10, 8, 10, 10, 8, 9, 44, 11, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_prep_sheet(wb, recipe_idx, recipes, sub_map, bom_lookup):
    if "Prep Sheet" in wb.sheetnames:
        del wb["Prep Sheet"]

    master_idx = wb.sheetnames.index("Master BOM")
    ws = wb.create_sheet(title="Prep Sheet", index=master_idx + 1)

    # ── Title row ──
    ws.cell(row=1, column=1, value="THE LARIAT — DAILY PREP SHEET")
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=16, color=DARK_GREEN)
    ws.merge_cells("A1:J1")

    # ── Multiplier input ──
    ws.cell(row=2, column=1, value="Date:")
    ws.cell(row=2, column=1).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=2, column=2, value=TODAY)
    ws.cell(row=2, column=2).font = Font(name="Arial", size=10, color="0000FF")

    ws.cell(row=2, column=4, value="Batch Multiplier →")
    ws.cell(row=2, column=4).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=2, column=6, value=1.0)
    ws.cell(row=2, column=6).font = Font(name="Arial", bold=True, size=12, color="0000FF")
    ws.cell(row=2, column=6).number_format = '0.0'
    ws.cell(row=2, column=6).alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=6).fill = PatternFill("solid", fgColor="FFFF00")

    dv = DataValidation(type="decimal", operator="greaterThan", formula1="0")
    dv.errorTitle = "Invalid"
    dv.error = "Enter a positive number"
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=2, column=6))

    ws.cell(row=2, column=8, value="(Blue = editable)")
    ws.cell(row=2, column=8).font = Font(name="Arial", italic=True, size=9, color="888888")

    ws.cell(row=3, column=1)  # blank row

    write_prep_headers(ws)

    # ── Checkbox data validation ──
    cb_dv = DataValidation(type="list", formula1='"✓"', allow_blank=True)
    cb_dv.errorTitle = "Check"
    cb_dv.error = "Use ✓ or leave blank"
    ws.add_data_validation(cb_dv)

    MULT_CELL = "$F$2"

    # ── Assign recipes to stations ──
    station_order = ["all", "grill", "fry", "saute", "salad", "expo"]
    station_recipes = {s: [] for s in station_order}

    for _, ridx in recipe_idx.iterrows():
        stations = normalize_station(ridx.get("station", ""))
        primary = stations[0] if stations[0] in station_order else "all"
        station_recipes[primary].append(ridx)

    r = 5
    group_ranges = []

    for station in station_order:
        rlist = station_recipes[station]
        if not rlist:
            continue

        # ── Station header ──
        color = STATION_COLORS.get(station, "6C757D")
        label = STATION_LABELS.get(station, station.upper())
        ws.cell(row=r, column=1, value="")
        ws.cell(row=r, column=2, value=f"━━━  {label}  ━━━")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
        for c in range(1, 13):
            ws.cell(row=r, column=c).font = Font(name="Arial", bold=True, size=12, color=WHITE)
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=color)
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center" if c == 2 else "left", vertical="center")
        ws.row_dimensions[r].height = 26
        r += 1

        sorted_rlist = sorted(rlist, key=lambda x: x["recipe_name"])

        for ridx_row in sorted_rlist:
            rid = ridx_row["recipe_id"]
            rname = ridx_row["recipe_name"]
            cat = ridx_row.get("category", "")
            yld = ridx_row.get("yield", "")
            yld_unit = ridx_row.get("yield_unit", "")
            notes_val = ridx_row.get("notes", "")

            recipe_row = r

            # ── Recipe header row ──
            ws.cell(row=r, column=1, value="")  # checkbox placeholder
            cb_dv.add(ws.cell(row=r, column=1))
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")

            ws.cell(row=r, column=2, value=rname)
            ws.cell(row=r, column=3, value=cat)
            ws.cell(row=r, column=8, value=yld if pd.notna(yld) else "")
            ws.cell(row=r, column=9, value=yld_unit if pd.notna(yld_unit) else "")
            ws.cell(row=r, column=10, value=safe_text(str(notes_val)) if pd.notna(notes_val) else "")
            ws.cell(row=r, column=10).alignment = Alignment(wrap_text=True)

            # Batch cost from BOM
            bom_data = bom_lookup.get(rid, {})
            batch_cost = sum(float(v["ext_cost_usd"]) for v in bom_data.values() if pd.notna(v.get("ext_cost_usd")))
            if batch_cost > 0:
                ws.cell(row=r, column=11, value=batch_cost)
                ws.cell(row=r, column=11).number_format = MONEY_FMT
                k_col = get_column_letter(11)
                ws.cell(row=r, column=12).value = f"={k_col}{r}*{MULT_CELL}"
                ws.cell(row=r, column=12).number_format = MONEY_FMT

            for c in range(1, 13):
                ws.cell(row=r, column=c).font = Font(name="Arial", bold=True, size=10, color=DARK_GREEN)
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
                ws.cell(row=r, column=c).border = Border(bottom=Side(style="medium", color=DARK_GREEN))
            r += 1
            group_start = r

            # ── Ingredient rows ──
            if rid in recipes:
                recipe_df = recipes[rid]
                for _, ing_row in recipe_df.iterrows():
                    ing = ing_row["ingredient"]
                    qty = ing_row.get("qty", "")
                    unit = ing_row.get("unit", "")
                    ing_notes = ing_row.get("notes", "")
                    is_sub = pd.notna(ing_notes) and "(sub-recipe)" in str(ing_notes)

                    if is_sub:
                        # ── Sub-recipe parent row ──
                        ws.cell(row=r, column=2, value=f"  ▸ {ing}")
                        ws.cell(row=r, column=4, value=qty if pd.notna(qty) else "")
                        ws.cell(row=r, column=5, value=unit if pd.notna(unit) else "")
                        if pd.notna(qty) and qty:
                            d_col = get_column_letter(4)
                            ws.cell(row=r, column=7).value = f"={d_col}{r}*{MULT_CELL}"
                            ws.cell(row=r, column=7).number_format = '0.00'

                        # Find matching sub-recipe
                        sub_rid = None
                        for sr in sub_map.get(rid, []):
                            if sr.replace("_", " ") in ing.replace("_", " ") or ing.replace("_", " ") in sr.replace("_", " "):
                                sub_rid = sr
                                break
                        if not sub_rid:
                            for sr in sub_map.get(rid, []):
                                if any(w in sr for w in ing.lower().split()):
                                    sub_rid = sr
                                    break

                        # Link to detail sheet
                        if sub_rid and sub_rid in recipes:
                            safe_name = sub_rid[:31]
                            if safe_name in [ws2.title for ws2 in wb.worksheets]:
                                ws.cell(row=r, column=2).hyperlink = f"#'{safe_name}'!A1"
                                ws.cell(row=r, column=2).font = Font(name="Arial", bold=True, italic=True, size=9, color="0563C1", underline="single")
                            else:
                                ws.cell(row=r, column=2).font = Font(name="Arial", bold=True, italic=True, size=9, color=MED_GREEN)
                        else:
                            ws.cell(row=r, column=2).font = Font(name="Arial", bold=True, italic=True, size=9, color=MED_GREEN)

                        for c in range(1, 13):
                            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=SUB_BG)
                            ws.cell(row=r, column=c).border = THIN_BORDER
                        r += 1

                        # ── Expanded sub-recipe ingredients ──
                        sub_group_start = r
                        if sub_rid and sub_rid in recipes:
                            sub_df = recipes[sub_rid]
                            for _, sub_ing in sub_df.iterrows():
                                si = sub_ing["ingredient"]
                                sq = sub_ing.get("qty", "")
                                su = sub_ing.get("unit", "")
                                sn = sub_ing.get("notes", "")

                                ws.cell(row=r, column=2, value=f"      · {si}")
                                ws.cell(row=r, column=4, value=sq if pd.notna(sq) else "")
                                ws.cell(row=r, column=5, value=su if pd.notna(su) else "")
                                if pd.notna(sq) and sq:
                                    d_col = get_column_letter(4)
                                    ws.cell(row=r, column=7).value = f"={d_col}{r}*{MULT_CELL}"
                                    ws.cell(row=r, column=7).number_format = '0.00'
                                ws.cell(row=r, column=10, value=safe_text(str(sn)) if pd.notna(sn) else "")

                                for c in range(1, 13):
                                    ws.cell(row=r, column=c).font = Font(name="Arial", size=9, color="555555")
                                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=SUB2_BG)
                                    ws.cell(row=r, column=c).border = THIN_BORDER
                                r += 1
                        if r > sub_group_start:
                            ws.row_dimensions.group(sub_group_start, r - 1, outline_level=2, hidden=False)
                    else:
                        # ── Regular ingredient row ──
                        ws.cell(row=r, column=2, value=f"  {ing}")
                        ws.cell(row=r, column=4, value=qty if pd.notna(qty) else "")
                        ws.cell(row=r, column=5, value=unit if pd.notna(unit) else "")
                        if pd.notna(qty) and qty:
                            d_col = get_column_letter(4)
                            ws.cell(row=r, column=7).value = f"={d_col}{r}*{MULT_CELL}"
                            ws.cell(row=r, column=7).number_format = '0.00'
                        ws.cell(row=r, column=10, value=safe_text(str(ing_notes)) if pd.notna(ing_notes) else "")

                        for c in range(1, 13):
                            ws.cell(row=r, column=c).font = Font(name="Arial", size=10)
                            ws.cell(row=r, column=c).border = THIN_BORDER
                        if (r - group_start) % 2 == 0:
                            for c in range(1, 13):
                                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                        r += 1

            if r > group_start:
                group_ranges.append((group_start, r - 1))

        # blank row between stations
        r += 1

    for start, end in group_ranges:
        ws.row_dimensions.group(start, end, outline_level=1, hidden=False)
    ws.sheet_properties.outlinePr.summaryBelow = False

    # ── Print setup ──
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    print(f"  Prep Sheet: {r} rows, {len(group_ranges)} recipe groups across {len([s for s in station_order if station_recipes[s]])} stations")


def main():
    print(f"Loading workbook: {WB_PATH.name}")
    wb = load_workbook(str(WB_PATH))
    print(f"  Existing sheets: {wb.sheetnames}")

    print("Loading canonical data...")
    recipe_idx, bom, recipes = load_data()
    sub_map = get_sub_recipes(recipe_idx)
    bom_lookup = build_bom_lookup(bom)

    print("Building Prep Sheet...")
    build_prep_sheet(wb, recipe_idx, recipes, sub_map, bom_lookup)

    print(f"Saving workbook...")
    wb.save(str(WB_PATH))
    print(f"Done — Prep Sheet added to {WB_PATH.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
