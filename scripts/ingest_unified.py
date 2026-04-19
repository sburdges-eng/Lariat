#!/usr/bin/env python3
"""Read Lariat_Unified_Workbook.xlsx (+ optional PDF) and print JSON for ingest.mjs."""
import json
import os
import re
import sys

import openpyxl

SOURCE = os.environ.get("LARIAT_SOURCE", "")
PDF_PATH = os.environ.get("LARIAT_PDF", "")

wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)


def line_items_from_sheet(ws):
    items = []
    started = False
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        a = str(row[0]).strip()
        al = a.lower()
        if al == "item":
            started = True
            continue
        if not started:
            continue
        if a and al != "item":
            items.append(a)
    return items


LINE_CHECK_SPECS = [
    ("brunch", ["BOH - Brunch Line Check", "BOH OOP - Brunch Line (working)"]),
    ("fry", ["BOH - Fry Line Check", "BOH OOP - Fry Line Check"]),
    ("garde", ["BOH - Garde Line Check", "BOH OOP - Line Check Garde"]),
    (
        "grille_saute",
        [
            "BOH - Grille-Saute Line Check",
            "BOH OOP - Line Check Grille_Sau",
            "BOH OOP - Line Check Grille_Saute",
        ],
    ),
]

line_checks = {}
for key, names in LINE_CHECK_SPECS:
    sn = next((n for n in names if n in wb.sheetnames), None)
    if not sn:
        line_checks[key] = []
        continue
    line_checks[key] = line_items_from_sheet(wb[sn])

setups = {}
if "BOH - Station Setup Procedures" in wb.sheetnames:
    ws = wb["BOH - Station Setup Procedures"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or not row[0]:
            continue
        station = str(row[0]).strip()
        step = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if step:
            setups.setdefault(station, []).append(step)

recipes = []
if "Recipe Book" in wb.sheetnames:
    ws = wb["Recipe Book"]
    current = None
    for row in ws.iter_rows(values_only=True):
        cells = (list(row) + [None] * 7)[:7]
        a, b, c, d, e, f, g = cells
        if a and f and "scale" in str(f).lower():
            if current:
                recipes.append(current)
            current = {"name": str(a).strip(), "ingredients": [], "procedure": []}
            continue
        if a and str(a).strip().lower() == "ingredient":
            continue
        if current is None:
            continue
        if a and b is not None:
            current["ingredients"].append(
                {
                    "item": str(a).strip(),
                    "qty": float(b) if isinstance(b, (int, float)) else b,
                    "unit": str(c).strip() if c else "",
                }
            )
        if g:
            current["procedure"].append(str(g).strip())
    if current:
        recipes.append(current)

allergen_kw = {
    "gluten": [
        "flour",
        "bread",
        "bun",
        "tortilla",
        "cornbread",
        "crouton",
        "batter",
        "panko",
        "sourdough",
        "waffle",
        "churro",
        "beer",
    ],
    "dairy": [
        "butter",
        "cheese",
        "milk",
        "cream",
        "buttermilk",
        "queso",
        "aioli",
        "yogurt",
        "mozzarella",
        "cotija",
        "parmesan",
    ],
    "egg": ["egg", "aioli", "mayo", "tartar"],
    "soy": ["soy", "miso", "tofu", "furikake"],
    "nuts": ["almond", "pecan", "walnut", "peanut", "pepita", "cashew", "hazelnut"],
    "fish": ["fish", "anchovy", "cod"],
    "shellfish": ["shrimp", "crab", "lobster", "oyster"],
    "sesame": ["sesame", "tahini", "furikake"],
}
for r in recipes:
    r["slug"] = re.sub(r"[^a-z0-9]+", "-", r["name"].lower()).strip("-")
    r["source"] = "excel"
    text = (r["name"] + " " + " ".join(i.get("item", "") for i in r.get("ingredients", []))).lower()
    r["allergens"] = sorted({a for a, kws in allergen_kw.items() if any(k in text for k in kws)})

staff = []
if "Labor - By Employee" in wb.sheetnames:
    ws = wb["Labor - By Employee"]
    seen = set()
    kitchen_kw = ("cook", "prep", "chef", "dish", "kitchen", "expo", "runner", "butcher", "km")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        last = str(row[0]).strip()
        first = str(row[1]).strip() if row[1] else ""
        if not first:
            continue
        job = str(row[2]).strip().lower() if row[2] else ""
        if not any(k in job for k in kitchen_kw):
            continue
        key = (last.lower(), first.lower())
        if key in seen:
            continue
        seen.add(key)
        sid = re.sub(r"[^a-z0-9]+", "_", (first + "_" + last).lower()).strip("_")
        staff.append(
            {
                "id": sid,
                "first": first,
                "last": last,
                "role": "cook",
                "active": True,
                "job_title": str(row[2]).strip() if row[2] else "",
            }
        )

pdf_added = 0
if PDF_PATH and os.path.exists(PDF_PATH):
    try:
        import pdfplumber

        MEASURE = re.compile(r"^\s*[\d¼½¾⅓⅔⅛⅜⅝⅞]")
        MEASURE_TOKENS = [
            "cup",
            "cups",
            "qt",
            "quart",
            "gallon",
            "lb",
            "lbs",
            "oz",
            "kg",
            "grams",
            "tsp",
            "tbsp",
            "ml",
            "case",
            "ea",
            "box",
            "bunch",
            "stick",
        ]

        def is_ing(line):
            l = line.strip().lower()
            if not l:
                return False
            if MEASURE.match(line):
                return True
            if l.startswith("●") or l.startswith("•") or l.startswith("-"):
                return True
            if " - " in l or " – " in l:
                rhs = re.split(r" [-–] ", l, 1)[-1]
                if MEASURE.match(rhs) or any(t in rhs for t in MEASURE_TOKENS):
                    return True
            for t in MEASURE_TOKENS:
                if re.search(rf"\b{re.escape(t)}\b", l):
                    return True
            return False

        def is_section(line):
            return line.strip().lower() in (
                "ingredients",
                "directions",
                "procedure",
                "method",
                "steps",
                "base",
                "soup base",
                "sauce",
            )

        with pdfplumber.open(PDF_PATH) as pdf:
            pages = [
                [l.rstrip() for l in (p.extract_text() or "").split("\n") if l.strip()]
                for p in pdf.pages
            ]
        pdf_recipes = []
        cur = None
        for i, lines in enumerate(pages):
            if i == 0 or not lines:
                continue
            first = lines[0].strip()
            is_header = (
                not is_ing(first)
                and not is_section(first)
                and 3 <= len(first) <= 60
                and not first.startswith("(")
            )
            if is_header:
                if cur:
                    pdf_recipes.append(cur)
                cur = {"name": first, "lines": lines[1:], "page": i + 1}
            elif cur:
                cur["lines"].extend(lines)
        if cur:
            pdf_recipes.append(cur)
        existing_slugs = {r["slug"] for r in recipes}
        existing_norm = {re.sub(r"[^a-z0-9]", "", r["name"].lower()) for r in recipes}
        for pr in pdf_recipes:
            slug = re.sub(r"[^a-z0-9]+", "-", pr["name"].lower()).strip("-")
            norm = re.sub(r"[^a-z0-9]", "", pr["name"].lower())
            if slug in existing_slugs or norm in existing_norm:
                continue
            ingredients = []
            procedure = []
            mode = "ing"
            for line in pr["lines"]:
                l = line.strip().lstrip("●•- ").strip()
                if not l:
                    continue
                ll = l.lower()
                if ll in ("ingredients", "base", "sauce", "seasoning"):
                    mode = "ing"
                    continue
                if ll in ("directions", "procedure", "method", "steps", "instructions"):
                    mode = "proc"
                    continue
                if re.match(r"^\d+[\.\)]\s", l):
                    mode = "proc"
                if mode == "ing":
                    ingredients.append({"item": l, "qty": None, "unit": ""})
                else:
                    procedure.append(l)
            text = (pr["name"] + " " + " ".join(i["item"] for i in ingredients)).lower()
            recipes.append(
                {
                    "name": pr["name"],
                    "slug": slug,
                    "source": "pdf",
                    "page": pr["page"],
                    "ingredients": ingredients,
                    "procedure": procedure,
                    "allergens": sorted(
                        {a for a, kws in allergen_kw.items() if any(k in text for k in kws)}
                    ),
                }
            )
            pdf_added += 1
    except ImportError:
        pass

out = {
    "_pdf_added": pdf_added,
    "line_checks": line_checks,
    "setups": setups,
    "recipes": recipes,
    "staff": staff,
}
print(json.dumps(out))
sys.stdout.flush()
