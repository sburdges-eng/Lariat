#!/usr/bin/env python3
"""
One-shot script: Update Lariat_Unified_Workbook.xlsx with:
  1. Navratil BEO invoice rows (📋 BEO Invoices)
  2. Navratil BEO prep rows (📋 BEO Prep)
  3. Replace old Pork Chop Marinade with Southwestern Rye version (Recipe Book)

Reads from EXPERIMENTS copy, writes dated output to same folder.
"""
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

SRC = Path.home() / "Desktop" / "LARIAT DEV" / "EXPERIMENTS" / "Lariat_Unified_Workbook.xlsx"
OUT = SRC.parent / f"Lariat_Unified_Workbook_{datetime.now():%Y-%m-%d}.xlsx"

# ── Navratil BEO Invoice data ──────────────────────────────────────────────
NAVRATIL_DATE = datetime(2026, 4, 10)
NAVRATIL_INVOICES = [
    # (Menu Item, Unit Cost, Qty, Line Total, Notes)
    ("Mini Rellenos",              4,   20,   80, None),
    ("Mac Balls",                  4,   20,   80, None),
    ("Battered Fish Taco",         7,   20,  140, None),
    ("Caprese Skewers",            4,   20,   80, None),
    ("Bri and Raspberry Philo Bites", 4, 20,  80, None),
    ("Caesar Salad Buffet",      150,    2,  300, None),
    ("Battered Fish Taco",       300,    1,  300, "Grilled"),
    ("Baked Ziti",               175,    1,  175, None),
    ("Prime Rib Carving Station",  35,  25,  875, None),
    ("Churros",                    4,   25,  100, None),
    ("Cupcakes",                  12,    5,   60, None),
    ("Trio Dips",                 30,    8,  240, None),
    ("Bar Spend Amount (?)",     570,    1,  570, None),
]
# Shared totals for this event
NAV_SUBTOTAL   = 3000
NAV_TAX        = 244.5
NAV_SVC_FEE    = 600
NAV_GRAND      = 3844.5
NAV_MIN_SPEND  = 3000
NAV_OVER_UNDER = 0

# ── Navratil BEO Prep data ────────────────────────────────────────────────
NAVRATIL_PREP = [
    # (Type, Item, Amount/Qty, Prep Day, Pre-Prep, Plating, Notes)
    ("Main Item", "Mini Rellenos",              20, None, None, "4 inch plastic", None),
    ("Main Item", "Mac Balls",                  20, None, None, "4 inch plastic", None),
    ("Main Item", "Battered Fish Taco",         20, None, None, None, None),
    ("Main Item", "Caprese Skewers",            20, None, None, "Platter", None),
    ("Main Item", "Bri and Raspberry Philo Bites", 20, None, None, "4 inch plastic", None),
    ("Main Item", "Caesar Salad Buffet",         2, None, None, "full pan", None),
    ("Main Item", "Battered Fish Taco",          1, None, None, "full pan", "grilled"),
    ("Main Item", "Baked Ziti",                  1, None, None, "full pan", None),
    ("Main Item", "Prime Rib Carving Station",  25, None, None, "Station", "Look into pricing or JP will buy"),
    ("Main Item", "Churros",                    25, None, None, "baskets", None),
    ("Main Item", "Cupcakes",                    5, None, None, None, "jp buys"),
    ("Main Item", "Trio Dips",                   8, None, None, "usual", None),
]

# ── New Southwestern Rye Pork Chop Marinade ───────────────────────────────
NEW_RECIPE_TITLE = "Southwestern Rye Pork Chop Marinade"
NEW_RECIPE_INGREDIENTS = [
    # (Ingredient, Base Qty, Base Unit, Display Unit, Procedure step)
    ("Rye Whiskey",              2,     "cup",  "cup",  "1. Combine all ingredients in a mixing bowl or cambro."),
    ("Worcestershire",           0.75,  "cup",  "cup",  "2. Whisk thoroughly until brown sugar is dissolved and oil is emulsified."),
    ("Dijon Mustard",            0.25,  "cup",  "cup",  "3. Yield: Approximately 3 QT of marinade."),
    ("Yellow Mustard",           0.25,  "cup",  "cup",  "4. Usage: Marinate pork chops for 6-12 hours before grilling."),
    ("Olive Oil",                0.25,  "cup",  "cup",  None),
    ("Apple Cider Vinegar",      2,     "tbsp", "tbsp", None),
    ("Brown Sugar",              0.25,  "cup",  "cup",  None),
    ("Shallots (finely minced)", 5,     "ea",   "ea",   None),
    ("Minced Garlic",            3,     "tbsp", "tbsp", None),
    ("Smoked Paprika",           1,     "tbsp", "tbsp", None),
    ("Ground Cumin",             1,     "tsp",  "tsp",  None),
    ("Coriander",                1,     "tsp",  "tsp",  None),
    ("White Pepper",             0.5,   "tsp",  "tsp",  None),
    ("Kosher Salt",              1,     "tbsp", "tbsp", None),
]


def copy_style(src_cell, dst_cell):
    """Copy formatting from one cell to another."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.alignment = copy(src_cell.alignment)


def make_adjusted_qty_formula(row_num, title_row):
    """Build the ArrayFormula for column E (Adjusted Qty)."""
    text = (
        f"=IF(C{row_num}=D{row_num},"
        f"B{row_num}*$G${title_row},"
        f"IFERROR(B{row_num}*$G${title_row}"
        f"*INDEX(UnitConversions[Factor],"
        f"MATCH(C{row_num}&\"|\"&D{row_num},"
        f"UnitConversions[From Unit]&\"|\"&UnitConversions[To Unit],0)),\"N/A\"))"
    )
    return ArrayFormula(ref=f"E{row_num}:E{row_num}", text=text)


def append_beo_invoices(wb):
    ws = wb["📋 BEO Invoices"]
    start = ws.max_row + 1
    # Copy style from last data row
    style_row = ws.max_row
    for i, (item, ucost, qty, total, notes) in enumerate(NAVRATIL_INVOICES):
        r = start + i
        values = [
            "Navratil", NAVRATIL_DATE, item, ucost, qty, total, notes,
            NAV_SUBTOTAL, NAV_TAX, NAV_SVC_FEE, NAV_GRAND, NAV_MIN_SPEND, NAV_OVER_UNDER,
        ]
        for c, val in enumerate(values, 1):
            dst = ws.cell(r, c, val)
            copy_style(ws.cell(style_row, c), dst)
    print(f"  BEO Invoices: appended {len(NAVRATIL_INVOICES)} rows (rows {start}-{start + len(NAVRATIL_INVOICES) - 1})")


def append_beo_prep(wb):
    ws = wb["📋 BEO Prep"]
    start = ws.max_row + 1
    style_row = ws.max_row
    for i, (typ, item, amt, prep_day, pre_prep, plating, notes) in enumerate(NAVRATIL_PREP):
        r = start + i
        values = ["Navratil", NAVRATIL_DATE, typ, item, amt, prep_day, pre_prep, plating, notes]
        for c, val in enumerate(values, 1):
            dst = ws.cell(r, c, val)
            copy_style(ws.cell(style_row, c), dst)
    print(f"  BEO Prep: appended {len(NAVRATIL_PREP)} rows (rows {start}-{start + len(NAVRATIL_PREP) - 1})")


def update_recipe_book(wb):
    ws = wb["Recipe Book"]

    # Find "Pork Chop Marinade" title row
    title_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value and "Pork Chop Marinade" in str(row[0].value):
            title_row = row[0].row
            break
    if title_row is None:
        print("  ERROR: Could not find 'Pork Chop Marinade' in Recipe Book")
        return False

    # Old recipe block: title, header, ingredients, blank separator
    # Find where the next recipe starts (next non-None in col A after blanks)
    old_end = title_row + 1  # at least header row
    in_blanks = False
    for r in range(title_row + 2, ws.max_row + 2):
        val = ws.cell(r, 1).value
        if val is None or str(val).strip() == "":
            in_blanks = True
            old_end = r
        elif in_blanks:
            # Hit next recipe
            break
        else:
            old_end = r

    old_count = old_end - title_row + 1  # rows used by old recipe (inclusive of trailing blanks)
    # New recipe: title(1) + header(1) + ingredients(14) + blanks(2) = 18 rows
    new_count = 1 + 1 + len(NEW_RECIPE_INGREDIENTS) + 2
    delta = new_count - old_count

    print(f"  Recipe Book: old recipe rows {title_row}-{old_end} ({old_count} rows)")
    print(f"  Recipe Book: new recipe needs {new_count} rows, delta={delta}")

    if delta > 0:
        ws.insert_rows(old_end + 1, delta)
        print(f"  Recipe Book: inserted {delta} rows after row {old_end}")
    elif delta < 0:
        ws.delete_rows(title_row + new_count, abs(delta))
        print(f"  Recipe Book: deleted {abs(delta)} excess rows")

    # Grab style from an ingredient row in a nearby recipe for consistent formatting
    # Use the first ingredient row of the old recipe (title_row + 2)
    style_src = title_row + 2

    # Write title row
    r = title_row
    ws.cell(r, 1, NEW_RECIPE_TITLE)
    for c in range(2, 7):
        ws.cell(r, c, None)
    ws.cell(r, 6, "Scale:")
    ws.cell(r, 7, 1)

    # Write header row
    r = title_row + 1
    for c, hdr in enumerate(["Ingredient", "Base Qty", "Base Unit", "Display Unit", "Adjusted Qty", None, "Procedure"], 1):
        ws.cell(r, c, hdr)

    # Write ingredient rows
    for i, (ingr, qty, bunit, dunit, proc) in enumerate(NEW_RECIPE_INGREDIENTS):
        r = title_row + 2 + i
        ws.cell(r, 1, ingr)
        ws.cell(r, 2, qty)
        ws.cell(r, 3, bunit)
        ws.cell(r, 4, dunit)
        ws.cell(r, 5).value = make_adjusted_qty_formula(r, title_row)
        ws.cell(r, 6, None)
        ws.cell(r, 7, proc)

    # Write blank separator rows
    for offset in range(len(NEW_RECIPE_INGREDIENTS), len(NEW_RECIPE_INGREDIENTS) + 2):
        r = title_row + 2 + offset
        for c in range(1, 8):
            ws.cell(r, c, None)

    print(f"  Recipe Book: wrote '{NEW_RECIPE_TITLE}' at rows {title_row}-{title_row + new_count - 1}")
    return True


def main() -> int:
    if not SRC.exists():
        print(f"ERROR: Source workbook not found: {SRC}")
        return 1

    print(f"Loading {SRC.name}...")
    wb = openpyxl.load_workbook(SRC)

    print("Updating sheets:")
    append_beo_invoices(wb)
    append_beo_prep(wb)
    ok = update_recipe_book(wb)
    if not ok:
        return 1

    wb.save(OUT)
    print(f"\nSaved to: {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
