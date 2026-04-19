#!/usr/bin/env python3
"""Build the Lariat Operations Workbook — hierarchical BOM with costing, order guide, batch scaling."""
import sys
from pathlib import Path
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

TODAY = date.today().isoformat()
OUT = ROOT / f"lariat_operations_workbook_{TODAY}.xlsx"

# ── Colors ──
DARK_GREEN = "1B4332"
MED_GREEN = "2D6A4F"
LIGHT_GREEN = "D8F3DC"
GOLD = "F4A261"
CREAM = "FFF8F0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
BLUE_INPUT = "0000FF"
RED_FLAG = "E63946"
SUB_BG = "EDF2F4"
SUB2_BG = "E2E8F0"

HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill("solid", fgColor=DARK_GREEN)
RECIPE_FONT = Font(name="Arial", bold=True, size=11, color=DARK_GREEN)
RECIPE_FILL = PatternFill("solid", fgColor=LIGHT_GREEN)
SUB_FONT = Font(name="Arial", bold=True, italic=True, size=10, color=MED_GREEN)
SUB_FILL = PatternFill("solid", fgColor=SUB_BG)
INGREDIENT_FONT = Font(name="Arial", size=10)
SUB_ING_FONT = Font(name="Arial", size=9, color="555555")
SUB_ING_FILL = PatternFill("solid", fgColor=SUB2_BG)
INPUT_FONT = Font(name="Arial", size=10, color=BLUE_INPUT)
MONEY_FMT = '$#,##0.00'
MONEY_FMT4 = '$#,##0.0000'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    bottom=Side(style="thin", color=MED_GRAY),
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
)
THICK_BOTTOM = Border(bottom=Side(style="medium", color=DARK_GREEN))


def load_data():
    recipe_idx = pd.read_csv(ROOT / "recipes" / "recipe_index.csv")
    bom = pd.read_csv(ROOT / "costing" / "bom_2026-04-05.csv")
    vendor_map = pd.read_csv(ROOT / "costing" / "ingredient_vendor_map.csv")
    recipes = {}
    norm_dir = ROOT / "recipes" / "normalized"
    for f in norm_dir.glob("*.csv"):
        rid = f.stem
        df = pd.read_csv(f)
        recipes[rid] = df
    fcr_path = ROOT / "costing" / "food_cost_report_2026-04-04.csv"
    fcr = pd.read_csv(fcr_path)
    return recipe_idx, bom, vendor_map, recipes, fcr


def get_sub_recipes(recipe_idx):
    subs = {}
    for _, row in recipe_idx.iterrows():
        sr = row.get("sub_recipes", "")
        if pd.notna(sr) and sr:
            subs[row["recipe_id"]] = [s.strip() for s in str(sr).split(";") if s.strip()]
        else:
            subs[row["recipe_id"]] = []
    return subs


def build_bom_lookup(bom):
    lookup = {}
    for _, row in bom.iterrows():
        rid = row["recipe_id"]
        if rid not in lookup:
            lookup[rid] = {}
        ing = row["ingredient"]
        lookup[rid][ing] = {
            "vendor": row.get("vendor", ""),
            "sku": row.get("sku", ""),
            "unit_price_usd": row.get("unit_price_usd", None),
            "pack_unit": row.get("pack_unit", ""),
            "pack_price_usd": row.get("pack_price_usd", None),
            "ext_cost_usd": row.get("ext_cost_usd", None),
            "costing_method": row.get("costing_method", ""),
        }
    return lookup


def safe_text(val):
    if pd.notna(val) and isinstance(val, str) and val.startswith("="):
        return "'" + val
    return val


def set_row_style(ws, row, font, fill=None, border=None, ncols=16):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = border


def money_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = MONEY_FMT
    return cell


def money4_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = MONEY_FMT4
    return cell


def pct_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = PCT_FMT
    return cell


def write_master_headers(ws):
    headers = [
        "Recipe / Ingredient", "Category", "Station", "Qty", "Unit",
        "Yield", "Yield Unit", "Portions/Batch",
        "Vendor", "SKU", "Unit Price", "Ext Cost",
        "Batch Cost", "Cost/Serving", "Menu Price", "Food Cost %"
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    widths = [32, 12, 10, 10, 8, 8, 10, 14, 10, 14, 12, 12, 12, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_master_sheet(wb, recipe_idx, recipes, bom_lookup, sub_map, fcr):
    ws = wb.active
    ws.title = "Master BOM"
    write_master_headers(ws)

    menu_links = {}
    for _, row in fcr.iterrows():
        rid = row.get("recipe_id")
        if pd.notna(rid) and rid:
            if rid not in menu_links:
                menu_links[rid] = {"price": row.get("price_usd"), "name": row.get("display_name")}

    sorted_recipes = recipe_idx.sort_values(["category", "recipe_name"])
    r = 2
    recipe_start_rows = {}
    group_ranges = []

    for _, ridx in sorted_recipes.iterrows():
        rid = ridx["recipe_id"]
        rname = ridx["recipe_name"]
        cat = ridx.get("category", "")
        station = ridx.get("station", "")
        yld = ridx.get("yield", "")
        yld_unit = ridx.get("yield_unit", "")

        recipe_start_rows[rid] = r
        recipe_row = r

        ws.cell(row=r, column=1, value=rname)
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=station)
        ws.cell(row=r, column=6, value=yld if pd.notna(yld) else "")
        ws.cell(row=r, column=7, value=yld_unit if pd.notna(yld_unit) else "")

        set_row_style(ws, r, RECIPE_FONT, RECIPE_FILL, THICK_BOTTOM)
        r += 1
        group_start = r

        if rid in recipes:
            recipe_df = recipes[rid]
            bom_data = bom_lookup.get(rid, {})

            for _, ing_row in recipe_df.iterrows():
                ing = ing_row["ingredient"]
                qty = ing_row.get("qty", "")
                unit = ing_row.get("unit", "")
                ppb = ing_row.get("portions_per_batch", "")
                notes = ing_row.get("notes", "")

                is_sub = pd.notna(notes) and "(sub-recipe)" in str(notes)

                if is_sub:
                    sub_rid = None
                    for sr in sub_map.get(rid, []):
                        if sr.replace("_", " ") in ing.replace("_", " ") or ing.replace("_", " ") in sr.replace("_", " "):
                            sub_rid = sr
                            break
                    if not sub_rid:
                        for sr in sub_map.get(rid, []):
                            if any(w in sr for w in ing.lower().split()):
                                sub_rid = sr
                                break

                    ws.cell(row=r, column=1, value=f"  ▸ {ing} (sub-recipe)")
                    ws.cell(row=r, column=4, value=qty if pd.notna(qty) else "")
                    ws.cell(row=r, column=5, value=unit if pd.notna(unit) else "")
                    ws.cell(row=r, column=8, value=str(ppb) if pd.notna(ppb) else "")
                    set_row_style(ws, r, SUB_FONT, SUB_FILL)

                    if sub_rid and sub_rid in recipes:
                        safe_name = sub_rid[:31]
                        ws.cell(row=r, column=1).hyperlink = f"#'{safe_name}'!A1"
                        ws.cell(row=r, column=1).font = Font(name="Arial", bold=True, italic=True, size=10, color="0563C1", underline="single")
                    r += 1

                    sub_group_start = r
                    if sub_rid and sub_rid in recipes:
                        sub_df = recipes[sub_rid]
                        sub_bom = bom_lookup.get(sub_rid, {})
                        for _, sub_ing in sub_df.iterrows():
                            si = sub_ing["ingredient"]
                            ws.cell(row=r, column=1, value=f"      · {si}")
                            ws.cell(row=r, column=4, value=sub_ing.get("qty", "") if pd.notna(sub_ing.get("qty")) else "")
                            ws.cell(row=r, column=5, value=sub_ing.get("unit", "") if pd.notna(sub_ing.get("unit")) else "")
                            sd = sub_bom.get(si, {})
                            ws.cell(row=r, column=9, value=sd.get("vendor", "") if pd.notna(sd.get("vendor")) else "")
                            ws.cell(row=r, column=10, value=str(sd.get("sku", "")) if pd.notna(sd.get("sku")) else "")
                            up = sd.get("unit_price_usd")
                            if pd.notna(up):
                                money4_cell(ws, r, 11, float(up))
                            ec = sd.get("ext_cost_usd")
                            if pd.notna(ec):
                                money_cell(ws, r, 12, float(ec))
                            set_row_style(ws, r, SUB_ING_FONT, SUB_ING_FILL, THIN_BORDER)
                            r += 1
                    if r > sub_group_start:
                        group_ranges.append((sub_group_start, r - 1, 2))
                else:
                    ws.cell(row=r, column=1, value=f"  {ing}")
                    ws.cell(row=r, column=4, value=qty if pd.notna(qty) else "")
                    ws.cell(row=r, column=5, value=unit if pd.notna(unit) else "")
                    ws.cell(row=r, column=8, value=str(ppb) if pd.notna(ppb) else "")
                    bd = bom_data.get(ing, {})
                    ws.cell(row=r, column=9, value=bd.get("vendor", "") if pd.notna(bd.get("vendor")) else "")
                    ws.cell(row=r, column=10, value=str(bd.get("sku", "")) if pd.notna(bd.get("sku")) else "")
                    up = bd.get("unit_price_usd")
                    if pd.notna(up):
                        money4_cell(ws, r, 11, float(up))
                    ec = bd.get("ext_cost_usd")
                    if pd.notna(ec):
                        money_cell(ws, r, 12, float(ec))
                    set_row_style(ws, r, INGREDIENT_FONT, None, THIN_BORDER)
                    r += 1

        if r > group_start:
            group_ranges.append((group_start, r - 1, 1))

        batch_cost_start = group_start
        batch_cost_end = r - 1
        if batch_cost_end >= batch_cost_start:
            l_col = get_column_letter(12)
            ws.cell(row=recipe_row, column=13).value = f"=SUM({l_col}{batch_cost_start}:{l_col}{batch_cost_end})"
            ws.cell(row=recipe_row, column=13).number_format = MONEY_FMT
            ws.cell(row=recipe_row, column=13).font = RECIPE_FONT

        ppb_val = ridx.get("yield", 1)
        if pd.notna(ppb_val) and float(ppb_val) > 0:
            m_col_letter = get_column_letter(13)
            ws.cell(row=recipe_row, column=14).value = f'=IF({m_col_letter}{recipe_row}>0,{m_col_letter}{recipe_row}/{float(ppb_val)},0)'
            ws.cell(row=recipe_row, column=14).number_format = MONEY_FMT
            ws.cell(row=recipe_row, column=14).font = RECIPE_FONT

        ml = menu_links.get(rid, {})
        mp = ml.get("price")
        if pd.notna(mp) if mp is not None else False:
            money_cell(ws, recipe_row, 15, float(mp))
            ws.cell(row=recipe_row, column=15).font = RECIPE_FONT
            n_col = get_column_letter(14)
            o_col = get_column_letter(15)
            ws.cell(row=recipe_row, column=16).value = f'=IF(AND({o_col}{recipe_row}>0,{n_col}{recipe_row}>0),{n_col}{recipe_row}/{o_col}{recipe_row},0)'
            ws.cell(row=recipe_row, column=16).number_format = PCT_FMT
            ws.cell(row=recipe_row, column=16).font = RECIPE_FONT

    for start, end, level in group_ranges:
        ws.row_dimensions.group(start, end, outline_level=level, hidden=False)

    ws.sheet_properties.outlinePr.summaryBelow = False
    return recipe_start_rows


def build_detail_sheets(wb, recipe_idx, recipes, bom_lookup, sub_map):
    subs_built = set()
    for _, ridx in recipe_idx.iterrows():
        rid = ridx["recipe_id"]
        for sr in sub_map.get(rid, []):
            if sr in subs_built or sr not in recipes:
                continue
            subs_built.add(sr)
            safe_name = sr[:31]
            ws = wb.create_sheet(title=safe_name)

            sr_idx = recipe_idx[recipe_idx["recipe_id"] == sr]
            sr_name = sr_idx.iloc[0]["recipe_name"] if len(sr_idx) > 0 else sr.replace("_", " ").title()

            ws.cell(row=1, column=1, value=sr_name)
            ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color=DARK_GREEN)
            ws.merge_cells("A1:H1")

            ws.cell(row=2, column=1, value="← Back to Master BOM")
            ws.cell(row=2, column=1).hyperlink = "#'Master BOM'!A1"
            ws.cell(row=2, column=1).font = Font(name="Arial", size=10, color="0563C1", underline="single")

            if len(sr_idx) > 0:
                ws.cell(row=3, column=1, value="Yield:")
                ws.cell(row=3, column=2, value=f"{sr_idx.iloc[0].get('yield', '')} {sr_idx.iloc[0].get('yield_unit', '')}")
                ws.cell(row=3, column=1).font = Font(name="Arial", bold=True, size=10)
                ws.cell(row=3, column=4, value="Station:")
                ws.cell(row=3, column=5, value=sr_idx.iloc[0].get("station", ""))
                ws.cell(row=3, column=4).font = Font(name="Arial", bold=True, size=10)

                notes_val = sr_idx.iloc[0].get("notes", "")
                if pd.notna(notes_val) and notes_val:
                    ws.cell(row=4, column=1, value="Notes:")
                    ws.cell(row=4, column=2, value=str(notes_val))
                    ws.cell(row=4, column=1).font = Font(name="Arial", bold=True, size=10)
                    ws.merge_cells("B4:H4")

            headers = ["Ingredient", "Qty", "Unit", "Vendor", "SKU", "Unit Price", "Ext Cost", "Notes"]
            hr = 6
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=hr, column=c, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")

            sub_df = recipes[sr]
            sub_bom = bom_lookup.get(sr, {})
            dr = hr + 1
            for _, ing_row in sub_df.iterrows():
                ing = ing_row["ingredient"]
                ws.cell(row=dr, column=1, value=ing)
                ws.cell(row=dr, column=2, value=ing_row.get("qty", "") if pd.notna(ing_row.get("qty")) else "")
                ws.cell(row=dr, column=3, value=ing_row.get("unit", "") if pd.notna(ing_row.get("unit")) else "")
                notes_val = ing_row.get("notes", "")
                ws.cell(row=dr, column=8, value=safe_text(str(notes_val)) if pd.notna(notes_val) else "")
                bd = sub_bom.get(ing, {})
                ws.cell(row=dr, column=4, value=bd.get("vendor", "") if pd.notna(bd.get("vendor")) else "")
                ws.cell(row=dr, column=5, value=str(bd.get("sku", "")) if pd.notna(bd.get("sku")) else "")
                up = bd.get("unit_price_usd")
                if pd.notna(up):
                    money4_cell(ws, dr, 6, float(up))
                ec = bd.get("ext_cost_usd")
                if pd.notna(ec):
                    money_cell(ws, dr, 7, float(ec))
                for c in range(1, 9):
                    ws.cell(row=dr, column=c).font = Font(name="Arial", size=10)
                    ws.cell(row=dr, column=c).border = THIN_BORDER
                dr += 1

            tc_col = get_column_letter(7)
            ws.cell(row=dr, column=6, value="Batch Total:")
            ws.cell(row=dr, column=6).font = Font(name="Arial", bold=True, size=10)
            ws.cell(row=dr, column=7).value = f"=SUM({tc_col}{hr+1}:{tc_col}{dr-1})"
            ws.cell(row=dr, column=7).number_format = MONEY_FMT
            ws.cell(row=dr, column=7).font = Font(name="Arial", bold=True, size=10)

            used_in = [ridx2["recipe_name"] for _, ridx2 in recipe_idx.iterrows()
                       if sr in sub_map.get(ridx2["recipe_id"], [])]
            if used_in:
                ur = dr + 2
                ws.cell(row=ur, column=1, value="Used in:")
                ws.cell(row=ur, column=1).font = Font(name="Arial", bold=True, size=10, color=MED_GREEN)
                for ui in used_in:
                    ur += 1
                    ws.cell(row=ur, column=1, value=f"  • {ui}")
                    ws.cell(row=ur, column=1).font = Font(name="Arial", size=10, color="555555")

            widths = [28, 10, 8, 10, 14, 12, 12, 30]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w


def build_order_guide(wb, recipe_idx, recipes, bom_lookup, sub_map):
    ws = wb.create_sheet(title="Order Guide")

    ws.cell(row=1, column=1, value="Lariat Order Guide & Batch Scaler")
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color=DARK_GREEN)
    ws.merge_cells("A1:G1")

    ws.cell(row=3, column=1, value="Batch Multiplier →")
    ws.cell(row=3, column=1).font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=3, column=2, value=1)
    ws.cell(row=3, column=2).font = INPUT_FONT
    ws.cell(row=3, column=2).number_format = '0.0'
    dv = DataValidation(type="decimal", operator="greaterThan", formula1="0")
    dv.errorTitle = "Invalid multiplier"
    dv.error = "Enter a positive number"
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=3, column=2))

    ws.cell(row=3, column=4, value="(Blue = editable inputs)")
    ws.cell(row=3, column=4).font = Font(name="Arial", italic=True, size=9, color="888888")

    headers = ["Ingredient", "Base Qty", "Unit", "Scaled Qty", "Vendor", "Unit Price", "Ext Cost"]
    hr = 5
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A6"

    agg = {}
    for _, ridx in recipe_idx.iterrows():
        rid = ridx["recipe_id"]
        if rid not in recipes:
            continue
        for _, ing_row in recipes[rid].iterrows():
            ing = ing_row["ingredient"]
            notes = ing_row.get("notes", "")
            if pd.notna(notes) and "(sub-recipe)" in str(notes):
                continue
            qty = ing_row.get("qty", 0)
            unit = ing_row.get("unit", "")
            if not pd.notna(qty):
                qty = 0
            key = (ing, unit)
            if key not in agg:
                bd = bom_lookup.get(rid, {}).get(ing, {})
                agg[key] = {
                    "total_qty": 0,
                    "vendor": bd.get("vendor", ""),
                    "sku": str(bd.get("sku", "")),
                    "unit_price": bd.get("unit_price_usd", None),
                    "recipes": [],
                }
            agg[key]["total_qty"] += float(qty) if pd.notna(qty) else 0
            agg[key]["recipes"].append(rid)

    r = hr + 1
    mult_cell = "$B$3"
    for (ing, unit), data in sorted(agg.items(), key=lambda x: x[0][0].lower()):
        ws.cell(row=r, column=1, value=ing)
        ws.cell(row=r, column=2, value=round(data["total_qty"], 2))
        ws.cell(row=r, column=3, value=unit)
        b_col = get_column_letter(2)
        ws.cell(row=r, column=4).value = f"={b_col}{r}*{mult_cell}"
        ws.cell(row=r, column=4).number_format = '0.00'
        ws.cell(row=r, column=4).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r, column=5, value=data["vendor"])
        up = data["unit_price"]
        if pd.notna(up) if up is not None else False:
            money4_cell(ws, r, 6, float(up))
            d_col = get_column_letter(4)
            f_col = get_column_letter(6)
            ws.cell(row=r, column=7).value = f"={d_col}{r}*{f_col}{r}"
            ws.cell(row=r, column=7).number_format = MONEY_FMT

        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            if not cell.font.bold:
                cell.font = Font(name="Arial", size=10)
            cell.border = THIN_BORDER
        r += 1

    total_row = r
    ws.cell(row=total_row, column=6, value="Total:")
    ws.cell(row=total_row, column=6).font = Font(name="Arial", bold=True, size=11)
    g_col = get_column_letter(7)
    ws.cell(row=total_row, column=7).value = f"=SUM({g_col}{hr+1}:{g_col}{total_row-1})"
    ws.cell(row=total_row, column=7).number_format = MONEY_FMT
    ws.cell(row=total_row, column=7).font = Font(name="Arial", bold=True, size=11)

    widths = [30, 12, 8, 12, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_boh_view(wb, recipe_idx, recipes, sub_map):
    ws = wb.create_sheet(title="BOH View")

    ws.cell(row=1, column=1, value="BOH Prep Reference (No Pricing)")
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color=DARK_GREEN)
    ws.merge_cells("A1:F1")

    headers = ["Recipe / Ingredient", "Category", "Station", "Qty", "Unit", "Notes"]
    hr = 3
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A4"

    sorted_recipes = recipe_idx.sort_values(["category", "recipe_name"])
    r = hr + 1
    group_ranges = []

    for _, ridx in sorted_recipes.iterrows():
        rid = ridx["recipe_id"]
        rname = ridx["recipe_name"]

        ws.cell(row=r, column=1, value=rname)
        ws.cell(row=r, column=2, value=ridx.get("category", ""))
        ws.cell(row=r, column=3, value=ridx.get("station", ""))
        yld = ridx.get("yield", "")
        yld_unit = ridx.get("yield_unit", "")
        ws.cell(row=r, column=6, value=f"Yield: {yld} {yld_unit}" if pd.notna(yld) else "")
        set_row_style(ws, r, RECIPE_FONT, RECIPE_FILL, THICK_BOTTOM, ncols=6)
        r += 1
        group_start = r

        if rid in recipes:
            for _, ing_row in recipes[rid].iterrows():
                ing = ing_row["ingredient"]
                notes = ing_row.get("notes", "")
                is_sub = pd.notna(notes) and "(sub-recipe)" in str(notes)

                prefix = "  ▸ " if is_sub else "  "
                ws.cell(row=r, column=1, value=f"{prefix}{ing}")
                ws.cell(row=r, column=4, value=ing_row.get("qty", "") if pd.notna(ing_row.get("qty")) else "")
                ws.cell(row=r, column=5, value=ing_row.get("unit", "") if pd.notna(ing_row.get("unit")) else "")
                ws.cell(row=r, column=6, value=safe_text(str(notes)) if pd.notna(notes) else "")

                if is_sub:
                    set_row_style(ws, r, SUB_FONT, SUB_FILL, THIN_BORDER, ncols=6)
                else:
                    set_row_style(ws, r, INGREDIENT_FONT, None, THIN_BORDER, ncols=6)
                r += 1

        if r > group_start:
            group_ranges.append((group_start, r - 1))

    for start, end in group_ranges:
        ws.row_dimensions.group(start, end, outline_level=1, hidden=False)
    ws.sheet_properties.outlinePr.summaryBelow = False

    widths = [34, 12, 10, 10, 8, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_summary_sheet(wb, recipe_idx, bom_lookup, recipes, fcr):
    ws = wb.create_sheet(title="Cost Summary")
    ws.cell(row=1, column=1, value="Recipe Cost Summary")
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color=DARK_GREEN)
    ws.merge_cells("A1:H1")

    headers = ["Recipe", "Category", "Yield", "Unit", "Batch Cost", "Cost/Serving", "Menu Price", "Food Cost %"]
    hr = 3
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A4"

    menu_links = {}
    for _, row in fcr.iterrows():
        rid = row.get("recipe_id")
        if pd.notna(rid) and rid:
            if rid not in menu_links:
                menu_links[rid] = float(row.get("price_usd", 0)) if pd.notna(row.get("price_usd")) else None

    sorted_recipes = recipe_idx.sort_values(["category", "recipe_name"])
    r = hr + 1
    for _, ridx in sorted_recipes.iterrows():
        rid = ridx["recipe_id"]
        ws.cell(row=r, column=1, value=ridx["recipe_name"])
        ws.cell(row=r, column=2, value=ridx.get("category", ""))
        yld = ridx.get("yield", 1)
        ws.cell(row=r, column=3, value=yld if pd.notna(yld) else "")
        ws.cell(row=r, column=4, value=ridx.get("yield_unit", ""))

        bom_data = bom_lookup.get(rid, {})
        batch_cost = sum(float(v["ext_cost_usd"]) for v in bom_data.values() if pd.notna(v.get("ext_cost_usd")))
        money_cell(ws, r, 5, batch_cost)

        if pd.notna(yld) and float(yld) > 0:
            e_col = get_column_letter(5)
            ws.cell(row=r, column=6).value = f"={e_col}{r}/{float(yld)}"
            ws.cell(row=r, column=6).number_format = MONEY_FMT

        mp = menu_links.get(rid)
        if mp:
            money_cell(ws, r, 7, mp)
            f_col = get_column_letter(6)
            g_col = get_column_letter(7)
            ws.cell(row=r, column=8).value = f'=IF({g_col}{r}>0,{f_col}{r}/{g_col}{r},"")'
            ws.cell(row=r, column=8).number_format = PCT_FMT

        for c in range(1, 9):
            ws.cell(row=r, column=c).font = Font(name="Arial", size=10)
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        r += 1

    widths = [28, 12, 8, 10, 12, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    print("Loading canonical data...")
    recipe_idx, bom, vendor_map, recipes, fcr = load_data()
    sub_map = get_sub_recipes(recipe_idx)
    bom_lookup = build_bom_lookup(bom)

    print(f"  {len(recipe_idx)} recipes, {len(recipes)} normalized, {len(bom)} BOM rows")
    print(f"  Sub-recipe map: {sum(1 for v in sub_map.values() if v)} recipes have sub-recipes")

    wb = Workbook()

    print("Building Master BOM sheet...")
    build_master_sheet(wb, recipe_idx, recipes, bom_lookup, sub_map, fcr)

    print("Building sub-recipe detail sheets...")
    build_detail_sheets(wb, recipe_idx, recipes, bom_lookup, sub_map)

    print("Building Cost Summary sheet...")
    build_summary_sheet(wb, recipe_idx, bom_lookup, recipes, fcr)

    print("Building Order Guide sheet...")
    build_order_guide(wb, recipe_idx, recipes, bom_lookup, sub_map)

    print("Building BOH View sheet...")
    build_boh_view(wb, recipe_idx, recipes, sub_map)

    print(f"Saving to {OUT}...")
    wb.save(str(OUT))
    print(f"Done — {OUT.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
