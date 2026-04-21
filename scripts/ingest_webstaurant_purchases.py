#!/usr/bin/env python3
"""Import WebstaurantStore spend-report purchases into the equipment table.

Parses the "Orders" sheet of one or more xlsx spend reports exported
from webstaurantstore.com and upserts each qualifying line as a row in
``equipment``. Item numbers (WebstaurantStore SKUs) become
``model_number``; order numbers become ``vendor_order_ref`` (prefixed
``#WS-``); ``vendor`` is always ``WebstaurantStore``.

Serial numbers, owner's manuals, and base warranty dates are *not* in
the export and must be entered by hand from the physical equipment.
However, WebstaurantStore's own Safeware extended-warranty line items
(SKU ``EXTWARN*``) appear in the same order as the equipment they
cover, with the coverage length in the product name (e.g.
``4 Year Extended Warranty Powered by Safeware``). When present, we
set ``warranty_expiration = purchase_date + N years`` on every
equipment row in that order and note the Safeware reference.

Usage::

    python3 scripts/ingest_webstaurant_purchases.py                     # import from data/imports/webstaurant/*.xlsx
    python3 scripts/ingest_webstaurant_purchases.py --include-smallwares
    python3 scripts/ingest_webstaurant_purchases.py --file path/to.xlsx --dry-run

Idempotent: uniqueness is ``(vendor, vendor_order_ref, model_number,
location_id)``; re-running skips existing rows.
"""
from __future__ import annotations

import argparse
import re
import sqlite3
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DB = ROOT / "data" / "lariat.db"
DEFAULT_IMPORT_DIR = ROOT / "data" / "imports" / "webstaurant"

EQUIPMENT_CATEGORIES = {"Restaurant Equipment", "Refrigeration"}
SMALLWARES_CATEGORIES = {"Tabletop", "Smallwares", "Storage & Transport"}
TOOL_CATEGORIES = {"Tools & Hardware"}

WARRANTY_SKU_PREFIX = "EXTWARN"
WARRANTY_YEARS_RE = re.compile(r"(\d+)\s*Year\s*Extended Warranty", re.I)
SAFEWARE_REF_RE = re.compile(r"Safeware\s+([A-Z]+:\d+)", re.I)


@dataclass
class OrderLine:
    purchase_date: str  # ISO
    order_number: int
    item_number: str
    description: str
    quantity: int
    wstore_category: str
    purchase_price: float
    user: str


def parse_money(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "")
    return float(s) if s else 0.0


def parse_date_iso(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def split_product(product: str) -> tuple[str, str]:
    """Return (item_number, description). Product is formatted
    ``"ITEMNUM - description..."``; if the dash is absent, the whole
    string is the description with item_number empty."""
    if not product:
        return "", ""
    parts = product.split(" - ", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return "", product.strip()


def classify(category_ws: str, description: str) -> str:
    """Map WebstaurantStore category + product text to a UI category."""
    d = (description or "").lower()
    if category_ws == "Refrigeration":
        return "Refrigeration"
    if category_ws == "Restaurant Equipment":
        if "fryer" in d:
            return "Fryers"
        if "oven" in d or "cook and hold" in d or "combi" in d:
            return "Ovens"
        if "mixer" in d or "slicer" in d or "processor" in d or "blender" in d:
            return "Prep & Mixers"
        return "Other"
    if category_ws in SMALLWARES_CATEGORIES:
        return "Smallwares"
    if category_ws in TOOL_CATEGORIES:
        return "Tools"
    return "Other"


def add_years_iso(iso: str, years: int) -> str | None:
    try:
        d = datetime.strptime(iso, "%Y-%m-%d")
    except (TypeError, ValueError):
        return None
    try:
        return d.replace(year=d.year + years).strftime("%Y-%m-%d")
    except ValueError:  # Feb 29 edge case
        return d.replace(year=d.year + years, day=28).strftime("%Y-%m-%d")


def iter_order_lines(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Orders" not in wb.sheetnames:
        raise SystemExit(f"{xlsx_path}: missing 'Orders' sheet (got {wb.sheetnames})")
    ws = wb["Orders"]
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = list(row)
            continue
        if not row or row[0] is None:
            continue
        d = {h: v for h, v in zip(header, row)}
        item_num, desc = split_product(d.get("Product") or "")
        iso = parse_date_iso(d.get("Date"))
        if iso is None:
            continue
        yield OrderLine(
            purchase_date=iso,
            order_number=int(d.get("Order Number") or 0),
            item_number=item_num,
            description=desc,
            quantity=int(d.get("Quantity") or 0),
            wstore_category=(d.get("Category") or "").strip(),
            purchase_price=parse_money(d.get("Purchase Price")),
            user=(d.get("User") or "").strip(),
        )


def collect_warranties(lines: list[OrderLine]) -> dict[int, tuple[int, str | None]]:
    """Return {order_number: (years, safeware_ref_or_None)} for any
    order whose line items include a Safeware extended warranty."""
    out: dict[int, tuple[int, str | None]] = {}
    for ln in lines:
        if not ln.item_number.upper().startswith(WARRANTY_SKU_PREFIX):
            continue
        m = WARRANTY_YEARS_RE.search(ln.description)
        if not m:
            continue
        years = int(m.group(1))
        ref_m = SAFEWARE_REF_RE.search(ln.description)
        ref = ref_m.group(1) if ref_m else None
        prev = out.get(ln.order_number)
        if prev is None or years > prev[0]:
            out[ln.order_number] = (years, ref)
    return out


def dedupe_lines(lines: list[OrderLine]) -> list[OrderLine]:
    """The full + recent spend reports overlap; dedupe by
    (order_number, item_number). Keep the first occurrence."""
    seen: set[tuple[int, str]] = set()
    out: list[OrderLine] = []
    for ln in lines:
        k = (ln.order_number, ln.item_number)
        if k in seen:
            continue
        seen.add(k)
        out.append(ln)
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--db", type=Path, default=DEFAULT_DB)
    ap.add_argument("--file", type=Path, action="append", default=None,
                    help="xlsx file(s) to ingest (repeatable). Defaults to every "
                         f"*.xlsx in {DEFAULT_IMPORT_DIR.relative_to(ROOT)}.")
    ap.add_argument("--location-id", default="default")
    ap.add_argument("--include-smallwares", action="store_true",
                    help="Also import Tabletop/Smallwares/Storage categories.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Parse and classify, print what would be inserted, no writes.")
    args = ap.parse_args()

    files: list[Path] = args.file or sorted(DEFAULT_IMPORT_DIR.glob("*.xlsx"))
    if not files:
        print(f"No xlsx files found in {DEFAULT_IMPORT_DIR}. Use --file to specify.",
              file=sys.stderr)
        return 2

    keep = set(EQUIPMENT_CATEGORIES) | set(TOOL_CATEGORIES)
    if args.include_smallwares:
        keep |= SMALLWARES_CATEGORIES

    all_lines: list[OrderLine] = []
    for f in files:
        print(f"Reading {f.relative_to(ROOT) if f.is_relative_to(ROOT) else f}")
        all_lines.extend(iter_order_lines(f))

    warranties = collect_warranties(all_lines)
    lines = dedupe_lines(all_lines)

    inserted = 0
    skipped_existing = 0
    skipped_category = 0
    skipped_warranty_line = 0
    warranty_applied = 0

    if not args.dry_run:
        con = sqlite3.connect(args.db)
        con.execute("PRAGMA foreign_keys=ON")
    else:
        con = None

    try:
        for ln in lines:
            if ln.item_number.upper().startswith(WARRANTY_SKU_PREFIX):
                skipped_warranty_line += 1
                continue
            if ln.wstore_category not in keep:
                skipped_category += 1
                continue
            if not ln.item_number or not ln.description:
                continue

            ui_cat = classify(ln.wstore_category, ln.description)
            order_ref = f"#WS-{ln.order_number}"

            warranty_iso: str | None = None
            warranty_note: str | None = None
            warr = warranties.get(ln.order_number)
            if warr is not None:
                years, ref = warr
                warranty_iso = add_years_iso(ln.purchase_date, years)
                warranty_note = (
                    f"Safeware {years}-year extended warranty"
                    + (f" ({ref})" if ref else "")
                    + f"; purchased with order {order_ref}."
                )
                warranty_applied += 1

            notes_parts: list[str] = []
            if ln.quantity and ln.quantity != 1:
                notes_parts.append(f"Quantity: {ln.quantity}.")
            if ln.user:
                notes_parts.append(f"Ordered by {ln.user}.")
            notes_parts.append(
                f"Imported from WebstaurantStore spend report "
                f"(category: {ln.wstore_category})."
            )
            if warranty_note:
                notes_parts.append(warranty_note)
            notes = " ".join(notes_parts)

            # Truncate very long names so the UI card stays readable;
            # the full description still lives in notes.
            name = ln.description if len(ln.description) <= 120 else ln.description[:117] + "…"

            if args.dry_run:
                print(f"  + [{ui_cat:12}] {ln.item_number:16} {order_ref:12} "
                      f"${ln.purchase_price:>9,.2f}  {name[:70]}"
                      + (f"  [warranty→{warranty_iso}]" if warranty_iso else ""))
                inserted += 1
                continue

            assert con is not None
            exists = con.execute(
                """SELECT id FROM equipment
                    WHERE vendor = ? AND vendor_order_ref = ?
                      AND model_number = ? AND location_id = ?""",
                ("WebstaurantStore", order_ref, ln.item_number, args.location_id),
            ).fetchone()
            if exists:
                skipped_existing += 1
                continue

            con.execute(
                """INSERT INTO equipment
                     (name, category, make_model, serial_number,
                      purchase_date, warranty_expiration, purchase_cost,
                      status, location_id, model_number, vendor,
                      vendor_order_ref, manual_path, notes)
                   VALUES (?, ?, NULL, NULL, ?, ?, ?, 'active', ?, ?, ?, ?, NULL, ?)""",
                (name, ui_cat, ln.purchase_date, warranty_iso,
                 ln.purchase_price, args.location_id,
                 ln.item_number, "WebstaurantStore", order_ref, notes),
            )
            inserted += 1

        if con is not None:
            con.commit()
    finally:
        if con is not None:
            con.close()

    verb = "would insert" if args.dry_run else "inserted"
    print(f"\nWebstaurantStore import summary:")
    print(f"  {verb:14}: {inserted}")
    print(f"  skipped (exists): {skipped_existing}")
    print(f"  skipped (other category): {skipped_category}")
    print(f"  warranty-line items skipped: {skipped_warranty_line}")
    print(f"  warranty dates applied: {warranty_applied}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
