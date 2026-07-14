#!/usr/bin/env python3
"""Update Master Costing workbook: confirmed ingredient maps + vendor price cross-refs."""
from __future__ import annotations

from pathlib import Path

import openpyxl

WORKBOOK = Path(__file__).resolve().parents[1] / "XL" / "Lariat_Master_Costing_2026-07-12.xlsx"

# Primary choice + Sysco/Shamrock competitor pairs (pack_price, unit_price per pack_unit).
VENDOR_ROWS: list[dict] = [
    {
        "ingredient": "SPICE, PEPPER BLK SHAKER GRIND PCH",
        "vendor": "shamrock",
        "sku": "4604681",
        "pack_size": 6,
        "pack_unit": "lb",
        "pack_price": 95.94,
        "unit_price": 15.99,
        "category": "Spices",
    },
    {
        "ingredient": "Spice Pepper Black Shaker Ground",
        "vendor": "sysco",
        "sku": "5661467",
        "pack_size": 15,
        "pack_unit": "lb",
        "pack_price": 298.77,
        "unit_price": 19.918,
        "category": "Spices (Sysco competitor)",
    },
    {
        "ingredient": "Sugar Granulated Extra Fine Cane",
        "vendor": "sysco",
        "sku": "4782694",
        "pack_size": 50,
        "pack_unit": "lb",
        "pack_price": 39.99,
        "unit_price": 0.7998,
        "category": "Canned & Dry",
    },
    {
        "ingredient": "SUGAR, WHT CANE GRANLTD X FINE BAG",
        "vendor": "shamrock",
        "sku": "3533721",
        "pack_size": 50,
        "pack_unit": "lb",
        "pack_price": 42.53,
        "unit_price": 0.8506,
        "category": "Canned & Dry (Shamrock competitor)",
    },
    {
        "ingredient": "JUICE, LEMON FRSH SQZ",
        "vendor": "shamrock",
        "sku": "2497471",
        "pack_size": 4,
        "pack_unit": "gal",
        "pack_price": 49.41,
        "unit_price": 12.3525,
        "category": "Produce",
    },
    {
        "ingredient": "Juice Lemon Pasteurized Ultra Premium",
        "vendor": "sysco",
        "sku": "4063095",
        "pack_size": 3,
        "pack_unit": "gal",
        "pack_price": 50.39,
        "unit_price": 16.796667,
        "category": "Produce (Sysco competitor)",
    },
    {
        "ingredient": "SAUCE, BFLO WING 1GAL JUG",
        "vendor": "shamrock",
        "sku": "3619071",
        "pack_size": 4,
        "pack_unit": "gal",
        "pack_price": 74.88,
        "unit_price": 18.72,
        "category": "Canned & Dry (Sweet Baby Ray's)",
    },
    {
        "ingredient": "Sauce Wing Buffalo",
        "vendor": "sysco",
        "sku": "3369347",
        "pack_size": 4,
        "pack_unit": "gal",
        "pack_price": 87.45,
        "unit_price": 21.8625,
        "category": "Canned & Dry (Sysco competitor SBR)",
    },
    {
        "ingredient": "PEPPER, CHILE HATCH 1/2\" GRN MILD DICED",
        "vendor": "shamrock",
        "sku": "3447731",
        "pack_size": 25,
        "pack_unit": "lb",
        "pack_price": 35.9,
        "unit_price": 1.436,
        "category": "Frozen (5/5 LB YGUNS)",
    },
    {
        "ingredient": "Chile Green Hatch Diced Mild",
        "vendor": "sysco",
        "sku": "7005009",
        "pack_size": 25,
        "pack_unit": "lb",
        "pack_price": 53.95,
        "unit_price": 2.158,
        "category": "Frozen (Sysco competitor)",
    },
    {
        "ingredient": "GARLIC, CLOVE PLD BAG REFRIG FRSH",
        "vendor": "shamrock",
        "sku": "2987171",
        "pack_size": 5,
        "pack_unit": "lb",
        "pack_price": 26.63,
        "unit_price": 5.326,
        "category": "Produce",
    },
    {
        "ingredient": "Garlic Peeled Fresh",
        "vendor": "sysco",
        "sku": "1821537",
        "pack_size": 20,
        "pack_unit": "lb",
        "pack_price": 90.95,
        "unit_price": 4.5475,
        "category": "Produce (Sysco competitor)",
    },
    {
        "ingredient": "Milk Whole Gallon",
        "vendor": "sysco",
        "sku": "4676306",
        "pack_size": 4,
        "pack_unit": "gal",
        "pack_price": 20.33,
        "unit_price": 5.0825,
        "category": "Dairy",
    },
    {
        "ingredient": "MILK, WHL 1GAL",
        "vendor": "shamrock",
        "sku": "4783491",
        "pack_size": 4,
        "pack_unit": "gal",
        "pack_price": 20.82,
        "unit_price": 5.205,
        "category": "Dairy (Shamrock competitor)",
    },
]

MAP_UPDATES: dict[str, tuple[str, str]] = {
    "pepper": ("SPICE, PEPPER BLK SHAKER GRIND PCH", "mapped"),
    "granulated sugar": ("Sugar Granulated Extra Fine Cane", "mapped"),
    "lemon juice": ("JUICE, LEMON FRSH SQZ", "mapped"),
    "franks hot sauce": ("SAUCE, BFLO WING 1GAL JUG", "mapped"),
    "hatch chile with juice": ("PEPPER, CHILE HATCH 1/2\" GRN MILD DICED", "mapped"),
    "hatch green chile": ("PEPPER, CHILE HATCH 1/2\" GRN MILD DICED", "mapped"),
    "green chile": ("PEPPER, CHILE HATCH 1/2\" GRN MILD DICED", "mapped"),
    "garlic puree": ("GARLIC, CLOVE PLD BAG REFRIG FRSH", "mapped"),
    "chopped garlic": ("GARLIC, CLOVE PLD BAG REFRIG FRSH", "mapped"),
    # Batch 2 — medium-confidence staples
    "salt": ("SALT, SEA WHT GRANULE 3LB KOSHER", "mapped"),
    "kosher salt": ("SALT, SEA WHT GRANULE 3LB KOSHER", "mapped"),
    "cumin": ("SPICE, CUMIN SEED GRND BULK", "mapped"),
    "paprika": ("SPICE, PAPRIKA SPANISH BULK", "mapped"),
    "cayenne": ("SPICE, PEPPER CAYENNE BULK", "mapped"),
    "ap flour": ("FLOUR, AP 50LB BAG", "mapped"),
    "butter": ("BUTTER, SOLID UNSLTD EURO STYL AA", "mapped"),
    "buttermilk": ("Buttermilk Whole Qt", "mapped"),
    "mayonnaise": ("MAYONNAISE, HVY DTY TUB", "mapped"),
    "honey": ("HONEY, LT AMBER GRD A 5LB JUG", "mapped"),
    "black pepper": ("SPICE, PEPPER BLK SHAKER GRIND PCH", "mapped"),
    "ground black pepper": ("SPICE, PEPPER BLK SHAKER GRIND PCH", "mapped"),
    "whole milk": ("Milk Whole Gallon", "mapped"),
    "milk": ("Milk Whole Gallon", "mapped"),
}


def norm_key(v) -> str:
    return str(v or "").strip().lower()


def upsert_vendor_row(ws, row_data: dict) -> str:
    key = (norm_key(row_data["ingredient"]), norm_key(row_data["vendor"]))
    for r in range(4, ws.max_row + 1):
        ing = ws.cell(r, 1).value
        vend = ws.cell(r, 2).value
        if norm_key(ing) == key[0] and norm_key(vend) == key[1]:
            for col, field in enumerate(
                ["ingredient", "vendor", "sku", "pack_size", "pack_unit", "pack_price", "unit_price", "category"],
                start=1,
            ):
                ws.cell(r, col, row_data[field])
            return "updated"
    next_row = ws.max_row + 1
    for col, field in enumerate(
        ["ingredient", "vendor", "sku", "pack_size", "pack_unit", "pack_price", "unit_price", "category"],
        start=1,
    ):
        ws.cell(next_row, col, row_data[field])
    return "added"


def upsert_map_row(ws, recipe_ingredient: str, vendor_ingredient: str, status: str) -> str:
    key = norm_key(recipe_ingredient)
    for r in range(4, ws.max_row + 1):
        if norm_key(ws.cell(r, 1).value) == key:
            ws.cell(r, 2, vendor_ingredient)
            ws.cell(r, 3, status)
            return "updated"
    next_row = ws.max_row + 1
    ws.cell(next_row, 1, recipe_ingredient)
    ws.cell(next_row, 2, vendor_ingredient)
    ws.cell(next_row, 3, status)
    return "added"


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK}")

    wb = openpyxl.load_workbook(WORKBOOK)
    vp = wb["Vendor Prices"]
    ivm = wb["Ingredient Vendor Map"]

    vp_results = [f"{upsert_vendor_row(vp, row)}: {row['ingredient']} ({row['vendor']})" for row in VENDOR_ROWS]
    map_results = [
        f"{upsert_map_row(ivm, recipe, vendor, status)}: {recipe} -> {vendor} [{status}]"
        for recipe, (vendor, status) in MAP_UPDATES.items()
    ]

    wb.save(WORKBOOK)
    print(f"Saved {WORKBOOK}")
    print("\nVendor Prices:")
    for line in vp_results:
        print(f"  {line}")
    print("\nIngredient Vendor Map:")
    for line in map_results:
        print(f"  {line}")


if __name__ == "__main__":
    main()
