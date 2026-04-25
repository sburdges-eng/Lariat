#!/usr/bin/env python3
"""Build Lariat Unified Workbook from originals + existing master + new CSVs.

Reads three input classes from the pre-scrub archive:

  1. `data/originals/` — raw supplier sources:
        - shamrock/Price List-2025.xls  (parsed from existing master workbook
          mirror; .xls parsing via pandas requires xlrd)
        - sysco/Sysco Purchase History.csv  (SUPC-indexed purchase rows)
        - sysco/BEO basics.csv              (BEO item catalog)
        - Toast/MenuItems.csv               (Toast menu item export)
        - webstaurantstore/total_spend_2026-04-18.csv
  2. `XL/Lariat Master Workbook.xlsx` — existing monolith with 80+ sheets,
     including the Shamrock Price List we mirror and the 68 Shamrock OC
     sheets (parsing deferred; see TODO block in build_shamrock_orders).
  3. `XL/table-2026-04-24-*.csv` — seven Toast analytics exports from
     2026-04-24 (one file is a duplicate of another; the duplicate is
     dropped before writing).
  4. `XL/Lariat Invoice *.xlsx` — seven BEO event invoice workbooks.

Writes `XL/Lariat_Unified_Workbook_2026-04-24.xlsx` with:
  - 🔧 Schema Documentation  (first sheet — sheet index + field defs)
  - 🔧 Master Product Catalog (LAR-XXXX IDs)
  - 🔧 Master Transaction Log (TXN-XXXX IDs; Shamrock purchases deferred)
  - 📋 BEO Invoices
  - 📈 Supplier Comparison   (Shamrock PL ⟷ Sysco PH by description)
  - 📊 Toast Analytics — Yearly Sales
  - 📊 Toast Analytics — Sales By Category
  - 📊 Toast Analytics — Labor Yearly
  - 📊 Toast Analytics — Labor Monthly
  - 📊 Toast Analytics — Labor Weekly
  - 📊 Toast Analytics — Menu Groups
  - ref — Shamrock Price List (carried from master)
  - ref — Sysco Purchase History
  - ref — Sysco BEO Basics
  - ref — Toast MenuItems
  - ref — Webstaurantstore Spend
  - 📊 Dashboard              (KPIs + 3 charts: yearly sales, top categories,
                                labor % monthly)
  - 📋 Deferred TODOs         (currently empty — all deferred items have landed)

Usage:
  python3 scripts/build_unified_workbook.py \
    --archive /Users/seanburdges/Dev/_archives/lariat-pre-scrub-2026-04-18 \
    --output  /Users/seanburdges/Dev/_archives/lariat-pre-scrub-2026-04-18/XL/Lariat_Unified_Workbook_2026-04-24.xlsx

  Both flags default to those exact paths.
"""
from __future__ import annotations

import argparse
import csv
import glob
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_ARCHIVE = "/Users/seanburdges/Dev/_archives/lariat-pre-scrub-2026-04-18"
DEFAULT_OUTPUT = (
    f"{DEFAULT_ARCHIVE}/XL/Lariat_Unified_Workbook_2026-04-24.xlsx"
)

CURRENCY_RE = re.compile(r"^\s*\$?-?[\d,]+(?:\.\d+)?\s*$")
HOURS_RE = re.compile(r"^\s*(\d+)h\s*(\d+)?m?\s*$")

# ── Parsing helpers ───────────────────────────────────────────────


def parse_currency(s: Any) -> float | None:
    """'$192,989' → 192989.0 ; '$0' → 0.0 ; '' / None → None."""
    if s is None:
        return None
    t = str(s).strip()
    if not t or t in {"-", "--"}:
        return None
    if not CURRENCY_RE.match(t):
        return None
    return float(t.replace("$", "").replace(",", ""))


def parse_int_loose(s: Any) -> int | None:
    """'5,743' → 5743 ; '24,536.0' → 24536 ; 'N' → None."""
    if s is None:
        return None
    t = str(s).strip().replace(",", "")
    if not t:
        return None
    try:
        return int(float(t))
    except ValueError:
        return None


def parse_hours(s: Any) -> float | None:
    """'5411h 33m' → 5411.55 (hours as float)."""
    if s is None:
        return None
    m = HOURS_RE.match(str(s))
    if not m:
        return None
    h = int(m.group(1))
    mins = int(m.group(2) or 0)
    return h + mins / 60.0


def parse_percent(s: Any) -> float | None:
    """'31.5%' → 31.5 ; '' → None."""
    if s is None:
        return None
    t = str(s).strip().rstrip("%")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


# ── Per-sheet builders ────────────────────────────────────────────


@dataclass
class SheetPayload:
    name: str
    header: list[str]
    rows: list[list[Any]] = field(default_factory=list)
    freeze_top: bool = True

    def row_count(self) -> int:
        return len(self.rows)


# Toast analytics: the seven CSVs land as six distinct sheets because
# 134853 duplicates 134741. We detect the dup by comparing file contents.
TOAST_CSV_MAPPING = [
    ("table-2026-04-24-133542.csv", "📊 Toast Analytics — Yearly Sales"),
    ("table-2026-04-24-133723.csv", "📊 Toast Analytics — Sales By Category"),
    ("table-2026-04-24-134036.csv", "📊 Toast Analytics — Labor Yearly"),
    ("table-2026-04-24-134047.csv", "📊 Toast Analytics — Labor Monthly"),
    ("table-2026-04-24-134554.csv", "📊 Toast Analytics — Labor Weekly"),
    ("table-2026-04-24-134741.csv", "📊 Toast Analytics — Menu Groups"),
    # 134853 dropped as duplicate (verified at runtime)
]


def build_toast_analytics(archive: Path) -> list[SheetPayload]:
    sheets: list[SheetPayload] = []
    # Dup guard: if 134741 and 134853 are byte-identical, skip 134853 as
    # documented. If they're different, we still skip 134853 but log a
    # warning — the caller decides whether to investigate.
    p741 = archive / "XL" / "table-2026-04-24-134741.csv"
    p853 = archive / "XL" / "table-2026-04-24-134853.csv"
    if p741.exists() and p853.exists():
        if p741.read_bytes() != p853.read_bytes():
            print(
                "[warn] 134741 and 134853 differ; only 134741 is being used. "
                "Review if menu-groups analytics came in two variants.",
                file=sys.stderr,
            )

    for fname, sheet_name in TOAST_CSV_MAPPING:
        path = archive / "XL" / fname
        if not path.exists():
            print(f"[skip] {fname} not found", file=sys.stderr)
            continue
        with path.open(newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh)
            rows = list(reader)
        if not rows:
            continue
        header = rows[0]
        body: list[list[Any]] = []
        for raw in rows[1:]:
            # Pad/truncate to header length and coerce obvious numerics.
            padded = (raw + [None] * len(header))[: len(header)]
            coerced: list[Any] = []
            for col, cell in zip(header, padded):
                if cell is None or cell == "":
                    coerced.append(None)
                    continue
                col_l = col.lower()
                if "hours" in col_l and col_l != "total hours" or col_l == "total hours":
                    parsed = parse_hours(cell)
                    coerced.append(parsed if parsed is not None else cell)
                elif (
                    "$" in str(cell)
                    or "sales" in col_l
                    or "cost" in col_l
                    or "splh" in col_l
                    or "discounts" in col_l
                    or "voids" in col_l
                    or "refunds" in col_l
                    or "tax" in col_l
                ):
                    parsed = parse_currency(cell)
                    coerced.append(parsed if parsed is not None else cell)
                elif "%" in str(cell):
                    parsed = parse_percent(cell)
                    coerced.append(parsed if parsed is not None else cell)
                else:
                    parsed = parse_int_loose(cell)
                    coerced.append(parsed if parsed is not None else cell)
            body.append(coerced)
        sheets.append(SheetPayload(sheet_name, header, body))
    return sheets


# ── Shamrock Price List (from existing Master Workbook) ──────────


def build_shamrock_price_list_ref(master_xlsx: Path) -> SheetPayload:
    wb = openpyxl.load_workbook(master_xlsx, data_only=True, read_only=True)
    try:
        ws = wb["shamrock - Price List-2025"]
    except KeyError:
        wb.close()
        return SheetPayload(
            "ref — Shamrock Price List",
            ["product_no", "description", "pack_size", "brand", "price", "unit"],
        )
    header = ["product_no", "description", "pack_size", "brand", "price", "unit"]
    rows: list[list[Any]] = []
    started = False
    for row in ws.iter_rows(values_only=True):
        # row layout: [blank, #, Product #, Description, Pack Size, Brand, Price, Unit]
        if not row or len(row) < 8:
            continue
        if not started:
            # Header row is the one where row[1] == '#' literally.
            if str(row[1]).strip() == "#":
                started = True
            continue
        product_no = str(row[2]).strip() if row[2] else None
        if not product_no:
            continue
        rows.append(
            [
                product_no,
                str(row[3]).strip() if row[3] else None,
                str(row[4]).strip() if row[4] else None,
                str(row[5]).strip() if row[5] else None,
                parse_currency(row[6]),
                str(row[7]).strip() if row[7] else None,
            ]
        )
    wb.close()
    return SheetPayload("ref — Shamrock Price List", header, rows)


# ── Shamrock Orders (from 68 OC sheets in Master Workbook) ─────────


SHAMROCK_ORDERS_HEADER = [
    "ship_date",
    "order_ref",
    "product_no",
    "description",
    "pack_size",
    "brand",
    "quantity",
    "price",
    "unit",
    "line_amount",
]


def parse_shamrock_oc_sheet(ws, sheet_name: str) -> list[list[Any]]:
    ship_date = None
    rows: list[list[Any]] = []
    started = False

    for row in ws.iter_rows(values_only=True):
        if not started:
            # Look for ship date
            for i, cell in enumerate(row):
                if cell and str(cell).strip().lower() == "ship date":
                    # find the next non-empty cell in this row
                    for j in range(i + 1, len(row)):
                        if row[j] is not None and str(row[j]).strip():
                            ship_date = str(row[j]).strip()
                            break
            # Check for header
            if len(row) > 2 and row[1] is not None and str(row[1]).strip() == "#":
                started = True
            continue

        if len(row) < 19:
            continue

        product_no = str(row[2]).strip() if row[2] else None
        if not product_no or product_no.lower() == "none":
            # stop if we hit an empty row after started
            if not any(row):
                break
            continue

        desc = str(row[3]).strip() if row[3] else None
        pack = str(row[9]).strip() if row[9] else None
        brand = str(row[10]).strip() if row[10] else None
        qty = parse_int_loose(row[13])
        price = parse_currency(row[14])
        unit = str(row[17]).strip() if row[17] else None
        amount = parse_currency(row[18])

        dt = ship_date
        if ship_date:
            m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", ship_date)
            if m:
                dt = f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"

        rows.append(
            [
                dt,
                sheet_name,
                product_no,
                desc,
                pack,
                brand,
                qty,
                price,
                unit,
                amount,
            ]
        )
    return rows


def build_shamrock_orders(master_xlsx: Path) -> SheetPayload:
    wb = openpyxl.load_workbook(master_xlsx, data_only=True, read_only=True)
    oc_sheet_names = [n for n in wb.sheetnames if n.startswith("Shamrock OC ")]
    all_rows: list[list[Any]] = []
    
    for sheet_name in oc_sheet_names:
        ws = wb[sheet_name]
        all_rows.extend(parse_shamrock_oc_sheet(ws, sheet_name))
        
    wb.close()
    
    # Sort by ship_date (descending) then order_ref
    all_rows.sort(key=lambda r: (r[0] or "", r[1] or ""), reverse=True)
    
    return SheetPayload("📋 Shamrock Orders", SHAMROCK_ORDERS_HEADER, all_rows)


# ── BEO Prep (mirror hand-curated sheet from Master Workbook) ────────


BEO_PREP_HEADER = [
    "client",
    "event_date",
    "type",
    "item",
    "amount_qty",
    "prep_day",
    "pre_prep",
    "plating",
]


def _format_event_date(v: Any) -> Any:
    """openpyxl returns datetime for the master sheet's date column;
    strings pass through unchanged."""
    if v is None:
        return None
    if hasattr(v, "date") and callable(v.date):
        return v.date().isoformat()
    if hasattr(v, "isoformat"):
        return v.isoformat()
    s = str(v).strip()
    return s or None


def build_beo_prep_ref(master_xlsx: Path) -> SheetPayload:
    """Mirror the hand-curated `📋 BEO Prep` sheet from the Master Workbook.

    The master sheet aggregates the per-event 'Kitchen Sheet' tabs in each
    Lariat Invoice xlsx and adds Type metadata (Main Item / Secondary Prep /
    Special Sauce) that the raw per-event sheets don't carry. Carried through
    verbatim — the master is source-of-truth.
    """
    wb = openpyxl.load_workbook(master_xlsx, data_only=True, read_only=True)
    try:
        ws = wb["📋 BEO Prep"]
    except KeyError:
        wb.close()
        return SheetPayload("📋 BEO Prep", BEO_PREP_HEADER)

    rows: list[list[Any]] = []
    started = False
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if not started:
            # Header row literally starts with 'Client'.
            if row[0] == "Client":
                started = True
            continue
        # Skip blank rows.
        if not any(v not in (None, "") for v in row):
            continue
        padded = list(row) + [None] * (8 - len(row))
        client = str(padded[0]).strip() if padded[0] else None
        event_date = _format_event_date(padded[1])
        type_ = str(padded[2]).strip() if padded[2] else None
        item = str(padded[3]).strip() if padded[3] else None
        amt = padded[4]
        if amt is not None and not isinstance(amt, (int, float)):
            amt = str(amt).strip() or None
        prep_day = str(padded[5]).strip() if padded[5] else None
        pre_prep = str(padded[6]).strip() if padded[6] else None
        plating = str(padded[7]).strip() if padded[7] else None
        rows.append(
            [client, event_date, type_, item, amt, prep_day, pre_prep, plating]
        )
    wb.close()
    return SheetPayload("📋 BEO Prep", BEO_PREP_HEADER, rows)


# ── Sysco Purchase History (line-item rows keyed by SUPC) ────────


SYSCO_PH_HEADER = [
    "supc",
    "pack",
    "size",
    "unit",
    "brand",
    "mfr_no",
    "description",
    "category",
    "case_price",
    "split_price",
    "per_lb",
    "net_wt_lb",
    "stock_status",
]


def _parse_sysco_flat_csv(path: Path) -> list[list[Any]]:
    """Sysco exports use H/F/P record markers. We only want P lines."""
    out: list[list[Any]] = []
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        for row in reader:
            if not row or row[0] != "P":
                continue
            # Columns per F-row: P, SUPC, CaseQty, SplitQty, Code, Status,
            # Replaced, Pack, Size, Unit, Brand, MfrNo, Desc, Cat, Case$,
            # Split$, PerLb, Market, Splittable, Splits, MinSplit, NetWt,
            # LeadTime, Stock, Substitute, Agr
            padded = (row + [None] * 27)[:27]
            out.append(
                [
                    padded[1],  # supc
                    padded[7],  # pack
                    padded[8],  # size
                    padded[9],  # unit
                    padded[10],  # brand
                    padded[11],  # mfr_no
                    padded[12],  # description
                    padded[13],  # category
                    parse_currency(padded[14]),  # case $
                    parse_currency(padded[15]),  # split $
                    str(padded[16]) if padded[16] is not None else None,  # per lb (Y/N)
                    parse_currency(padded[21]) if padded[21] is not None else None,
                    padded[23],  # stock status
                ]
            )
    return out


def build_sysco_purchase_history_ref(archive: Path) -> SheetPayload:
    path = archive / "data" / "originals" / "sysco" / "Sysco Purchase History.csv"
    if not path.exists():
        return SheetPayload("ref — Sysco Purchase History", SYSCO_PH_HEADER)
    return SheetPayload(
        "ref — Sysco Purchase History", SYSCO_PH_HEADER, _parse_sysco_flat_csv(path)
    )


def build_sysco_beo_basics_ref(archive: Path) -> SheetPayload:
    path = archive / "data" / "originals" / "sysco" / "BEO basics.csv"
    if not path.exists():
        return SheetPayload("ref — Sysco BEO Basics", SYSCO_PH_HEADER)
    return SheetPayload(
        "ref — Sysco BEO Basics", SYSCO_PH_HEADER, _parse_sysco_flat_csv(path)
    )


# ── Toast MenuItems ──────────────────────────────────────────────


def build_toast_menu_items_ref(archive: Path) -> SheetPayload:
    path = archive / "data" / "originals" / "Toast" / "MenuItems.csv"
    header = [
        "item_id",
        "guid",
        "name",
        "number",
        "imported_id",
        "base_price",
        "created_date",
        "archived",
        "modifier",
        "sku",
        "plu",
    ]
    if not path.exists():
        return SheetPayload("ref — Toast MenuItems", header)
    rows: list[list[Any]] = []
    # Toast exports mix UTF-8 and Latin-1 bytes in practice; tolerate.
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        reader = csv.DictReader(fh)
        for r in reader:
            rows.append(
                [
                    r.get("Item ID") or None,
                    r.get("GUID") or None,
                    r.get("Name") or None,
                    r.get("Number") or None,
                    r.get("Imported ID") or None,
                    parse_currency(r.get("Base Price")),
                    r.get("Created Date") or None,
                    r.get("Archived") or None,
                    r.get("Modifier") or None,
                    r.get("SKU") or None,
                    r.get("PLU") or None,
                ]
            )
    return SheetPayload("ref — Toast MenuItems", header, rows)


# ── Webstaurantstore total spend ─────────────────────────────────


def build_webstaurantstore_ref(archive: Path) -> SheetPayload:
    path = (
        archive
        / "data"
        / "originals"
        / "webstaurantstore"
        / "total_spend_2026-04-18.csv"
    )
    header = ["total_spend", "shipping_address"]
    if not path.exists():
        return SheetPayload("ref — Webstaurantstore Spend", header)
    # Format is non-standard: first line has a "Grand Total: " cell then the
    # actual fields. We just scan for a data row matching the header layout.
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = list(csv.reader(fh))
    rows: list[list[Any]] = []
    for row in reader[2:]:  # skip two-row header
        if not row or not row[0].strip():
            continue
        total = parse_currency(row[0])
        addr = row[1] if len(row) > 1 else None
        rows.append([total, addr])
    return SheetPayload("ref — Webstaurantstore Spend", header, rows)


# ── BEO invoices (parse the 7 event xlsx files) ──────────────────


BEO_INVOICES_HEADER = [
    "event_file",
    "client",
    "item",
    "cost_per_unit",
    "quantity",
    "line_total",
    "notes",
]


def _extract_client_from_filename(stem: str) -> str:
    # "Lariat Invoice Christy Nichols 9_7 " → "Christy Nichols"
    # "Invoice Darrell and anne collett 9_27" → "Darrell and anne collett"
    s = stem.replace("Lariat Invoice", "").replace("Invoice", "").strip()
    # Strip trailing date-like tokens (1-2 digits + _ + digits)
    s = re.sub(r"\s+\d{1,2}[_.\- ]\d{1,2}(?:[_.\- ]\d{2,4})?\s*$", "", s)
    return s.strip() or stem


def build_beo_invoices(archive: Path) -> SheetPayload:
    pattern = str(archive / "XL" / "Lariat Invoice *.xlsx")
    paths = sorted(glob.glob(pattern))
    # Also pick up the Darrell invoice (filename differs).
    paths += sorted(glob.glob(str(archive / "XL" / "Invoice *.xlsx")))
    rows: list[list[Any]] = []
    for p in paths:
        client = _extract_client_from_filename(Path(p).stem)
        try:
            wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        except Exception as err:
            print(f"[beo] failed to open {p}: {err}", file=sys.stderr)
            continue
        ws = wb[wb.sheetnames[0]]
        started = False
        for row in ws.iter_rows(values_only=True):
            # row[0..4] = Item, Cost, Amount, Total, Notes (left block)
            if not row or len(row) < 4:
                continue
            if not started:
                if row[0] and str(row[0]).strip().lower() == "item":
                    started = True
                continue
            item = str(row[0]).strip() if row[0] else None
            if not item:
                # blank row ends the left block; stop reading this sheet.
                break
            cost = parse_currency(row[1])
            qty = parse_currency(row[2])  # may be float like 20.0
            total = parse_currency(row[3])
            notes = str(row[4]).strip() if len(row) > 4 and row[4] else None
            rows.append([Path(p).name, client, item, cost, qty, total, notes])
        wb.close()
    return SheetPayload("📋 BEO Invoices", BEO_INVOICES_HEADER, rows)


# ── Derived: Master Product Catalog ──────────────────────────────


MPC_HEADER = [
    "lariat_id",
    "supplier",
    "supplier_product_no",
    "description",
    "pack_size",
    "brand",
    "price",
    "unit",
    "category",
    "source_list",
]


def build_master_product_catalog(
    shamrock: SheetPayload, sysco_ph: SheetPayload, sysco_beo: SheetPayload
) -> SheetPayload:
    """Unify Shamrock PL + Sysco PH + Sysco BEO into LAR-XXXX-keyed rows.

    Dedup key: (supplier, supplier_product_no). Collisions across suppliers
    get distinct LAR IDs. Price is whichever row won (first-seen wins).
    """
    seen: set[tuple[str, str]] = set()
    next_id = 1001
    rows: list[list[Any]] = []

    def emit(supplier: str, product_no: Any, description: Any, pack: Any,
             brand: Any, price: Any, unit: Any, category: Any, source: str) -> None:
        nonlocal next_id
        pno = (product_no or "").strip() if isinstance(product_no, str) else str(product_no or "").strip()
        if not pno:
            return
        key = (supplier, pno)
        if key in seen:
            return
        seen.add(key)
        rows.append(
            [
                f"LAR-{next_id}",
                supplier,
                pno,
                description,
                pack,
                brand,
                price,
                unit,
                category,
                source,
            ]
        )
        next_id += 1

    # Shamrock: (product_no, description, pack_size, brand, price, unit)
    for r in shamrock.rows:
        emit("Shamrock", r[0], r[1], r[2], r[3], r[4], r[5], None, "Price List-2025")

    # Sysco PH: (supc, pack, size, unit, brand, mfr, desc, cat, case$, split$, …)
    for src_name, payload in [
        ("Sysco Purchase History", sysco_ph),
        ("Sysco BEO Basics", sysco_beo),
    ]:
        for r in payload.rows:
            pack_size = f"{r[1]}/{r[2]}/{r[3]}" if r[1] and r[2] and r[3] else None
            emit(
                "Sysco",
                r[0],
                r[6],
                pack_size,
                r[4],
                r[8],  # case price
                r[3],  # unit
                r[7],  # category
                src_name,
            )

    return SheetPayload("🔧 Master Product Catalog", MPC_HEADER, rows)


# ── Derived: Supplier Comparison ─────────────────────────────────


SC_HEADER = [
    "description_normalized",
    "shamrock_product_no",
    "shamrock_description",
    "shamrock_price",
    "shamrock_unit",
    "sysco_supc",
    "sysco_description",
    "sysco_case_price",
    "sysco_unit",
]


def _normalize_desc(s: Any) -> str:
    if s is None:
        return ""
    t = str(s).upper()
    # Drop noisy punctuation, collapse spaces.
    t = re.sub(r"[^A-Z0-9 ]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def build_supplier_comparison(
    shamrock: SheetPayload, sysco_ph: SheetPayload
) -> SheetPayload:
    """Side-by-side rows where normalized first-word bucket matches.

    Not fuzzy: we bucket by the first token (e.g. 'BEEF', 'CHICKEN') and
    emit cartesian pairs where descriptions share the first *two* tokens.
    True fuzzy matching is out of scope; this gives operators a rough
    side-by-side to skim.
    """
    # Build bucket map for Sysco by first 2 tokens.
    sysco_by_prefix: dict[str, list[list[Any]]] = {}
    for r in sysco_ph.rows:
        n = _normalize_desc(r[6])
        if not n:
            continue
        toks = n.split()
        prefix = " ".join(toks[:2]) if len(toks) >= 2 else (toks[0] if toks else "")
        sysco_by_prefix.setdefault(prefix, []).append(r)

    out: list[list[Any]] = []
    for sr in shamrock.rows:
        n = _normalize_desc(sr[1])
        if not n:
            continue
        toks = n.split()
        prefix = " ".join(toks[:2]) if len(toks) >= 2 else (toks[0] if toks else "")
        matches = sysco_by_prefix.get(prefix, [])
        if not matches:
            # Still emit the Shamrock row alone so the comparison sheet
            # is a full left-join view (operators want to see "we buy
            # this from Shamrock; no Sysco equivalent available").
            out.append(
                [n, sr[0], sr[1], sr[4], sr[5], None, None, None, None]
            )
            continue
        for m in matches:
            out.append(
                [
                    n,
                    sr[0],
                    sr[1],
                    sr[4],
                    sr[5],
                    m[0],  # supc
                    m[6],  # sysco desc
                    m[8],  # sysco case price
                    m[3],  # sysco unit
                ]
            )
    # Sort by normalized description so the sheet reads alphabetically.
    out.sort(key=lambda r: r[0])
    return SheetPayload("📈 Supplier Comparison", SC_HEADER, out)


# ── Derived: Master Transaction Log (BEO revenue only, purchases TODO) ──


MTL_HEADER = [
    "transaction_id",
    "date",
    "type",
    "source_counterparty",
    "product_no",
    "description",
    "qty",
    "unit_price",
    "unit",
    "total_amount",
    "reference_no",
    "status",
]


def build_master_transaction_log(beo: SheetPayload, shamrock_orders: SheetPayload) -> SheetPayload:
    """BEO rows as revenue transactions and Shamrock purchases as expense transactions."""
    rows: list[list[Any]] = []
    next_id = 5001
    # Try to pluck the event date from the filename (e.g. 'Navratil 4_10.xlsx').
    file_date_re = re.compile(r"(\d{1,2})[_.\- ](\d{1,2})")
    for r in beo.rows:
        event_file, client, item, cost, qty, total, _notes = r
        m = file_date_re.search(event_file or "")
        if m:
            month, day = int(m.group(1)), int(m.group(2))
            date = f"2025-{month:02d}-{day:02d}"  # BEO files are all 2025 events
        else:
            date = None
        rows.append(
            [
                f"TXN-{next_id}",
                date,
                "BEO Revenue",
                f"Client: {client}" if client else None,
                None,
                item,
                qty,
                cost,
                None,
                total,
                f"BEO-{Path(event_file).stem}" if event_file else None,
                "Invoiced",
            ]
        )
        next_id += 1

    for r in shamrock_orders.rows:
        ship_date, order_ref, product_no, description, pack_size, brand, quantity, price, unit, line_amount = r
        rows.append(
            [
                f"TXN-{next_id}",
                ship_date,
                "Purchase",
                "Vendor: Shamrock",
                product_no,
                description,
                quantity,
                price,
                unit,
                line_amount,
                order_ref,
                "Delivered",
            ]
        )
        next_id += 1

    # Sort by date (descending)
    rows.sort(key=lambda r: r[1] or "", reverse=True)

    return SheetPayload("🔧 Master Transaction Log", MTL_HEADER, rows)


# ── Derived: Pricing Trends ──────────────────────────────────────


PRICING_TRENDS_HEADER = [
    "product_no",
    "description",
    "first_order_date",
    "last_order_date",
    "order_count",
    "min_price",
    "max_price",
    "first_price",
    "last_price",
    "pct_change",
]


def build_pricing_trends(shamrock_orders: SheetPayload) -> SheetPayload:
    grouped: dict[tuple[str, str], list[tuple[str, float]]] = {}
    for r in shamrock_orders.rows:
        ship_date, order_ref, product_no, description, pack_size, brand, quantity, price, unit, line_amount = r
        if not product_no or price is None or not ship_date:
            continue
        key = (product_no, str(description or ""))
        grouped.setdefault(key, []).append((ship_date, float(price)))

    out_rows: list[list[Any]] = []
    for (product_no, description), records in grouped.items():
        if len(records) < 2:
            continue
        records.sort(key=lambda x: x[0])
        first_date, first_price = records[0]
        last_date, last_price = records[-1]

        prices = [p for _, p in records]
        min_price = min(prices)
        max_price = max(prices)

        pct_change = None
        if first_price > 0:
            pct_change = (last_price - first_price) / first_price

        out_rows.append(
            [
                product_no,
                description,
                first_date,
                last_date,
                len(records),
                min_price,
                max_price,
                first_price,
                last_price,
                pct_change,
            ]
        )

    out_rows.sort(key=lambda r: r[9] if r[9] is not None else 0, reverse=True)
    return SheetPayload("📈 Pricing Trends", PRICING_TRENDS_HEADER, out_rows)


# ── Schema Documentation sheet ──────────────────────────────────


def build_schema_documentation(counts: dict[str, int]) -> SheetPayload:
    """Inline schema-doc sheet summarizing what this run actually produced.

    Intentionally differs from the old Schema Documentation.csv: that file
    was written by hand in April to describe an *aspirational* state. This
    sheet describes what THIS build actually emitted, with live row counts.
    """
    header = ["field_0", "field_1", "field_2", "field_3", "field_4"]
    rows: list[list[Any]] = [
        ["LARIAT UNIFIED WORKBOOK 2026-04-24", "", "", "", ""],
        ["Generated by scripts/build_unified_workbook.py", "", "", "", ""],
        ["", "", "", "", ""],
        ["SHEET INDEX", "", "", "", ""],
        ["Sheet", "Purpose", "Rows", "", ""],
    ]
    for name, n in counts.items():
        rows.append([name, "", n, "", ""])
    rows.extend(
        [
            ["", "", "", "", ""],
            ["FIELD DEFINITIONS — MASTER PRODUCT CATALOG", "", "", "", ""],
            ["Field", "Type", "Format", "Nullable", "Example"],
            ["lariat_id", "String", "LAR-XXXX", "No", "LAR-1001"],
            ["supplier", "String", "free text", "No", "Shamrock / Sysco"],
            ["supplier_product_no", "String", "alphanumeric", "No", "4608121"],
            ["description", "String", "free text", "No", "BEEF, CHEEK MEAT REFRIG"],
            ["pack_size", "String", "X/Y/UNIT", "Yes", "1/5/LB"],
            ["brand", "String", "abbreviation", "Yes", "KATYS"],
            ["price", "Currency", "numeric", "Yes", "57.26"],
            ["unit", "String", "CS/EA/LB", "No", "CS"],
            ["category", "String", "free text", "Yes", "Meat-Beef"],
            ["source_list", "String", "free text", "Yes", "Price List-2025"],
            ["", "", "", "", ""],
            ["FIELD DEFINITIONS — MASTER TRANSACTION LOG", "", "", "", ""],
            ["Field", "Type", "Format", "Nullable", "Example"],
            ["transaction_id", "String", "TXN-XXXX", "No", "TXN-5001"],
            ["date", "Date", "YYYY-MM-DD", "Yes", "2025-09-27"],
            ["type", "Enum", "Purchase | BEO Revenue", "No", "BEO Revenue"],
            ["source_counterparty", "String", "free text", "Yes", "Client: Darrell"],
            ["product_no", "String", "alphanumeric", "Yes", "4608121"],
            ["description", "String", "free text", "No", "Mac Balls"],
            ["qty", "Number", "numeric", "Yes", "20"],
            ["unit_price", "Currency", "numeric", "Yes", "4.00"],
            ["unit", "String", "CS/EA/LB", "Yes", ""],
            ["total_amount", "Currency", "numeric", "Yes", "80.00"],
            ["reference_no", "String", "BEO-… / OC-…", "Yes", "BEO-Navratil 4_10"],
            ["status", "String", "free text", "Yes", "Invoiced"],
        ]
    )
    return SheetPayload(
        "🔧 Schema Documentation", header, rows, freeze_top=False
    )


# ── Deferred TODOs sheet ─────────────────────────────────────────


def build_deferred_todos() -> SheetPayload:
    header = ["deferred_sheet", "blocker", "next_step"]
    rows: list[list[Any]] = []  # all previously-deferred items now landed
    return SheetPayload("📋 Deferred TODOs", header, rows)


# ── Dashboard (KPIs + charts) ────────────────────────────────────


@dataclass
class DashboardData:
    """Inputs for the dashboard sheet — passed in already-built so the
    dashboard builder doesn't reach back into archive paths."""
    yearly_sales: SheetPayload    # cols: year, net_sales, gross_sales, orders, checks, guests, ...
    sales_by_cat: SheetPayload    # cols: year, sales_category, net_sales
    labor_monthly: SheetPayload   # cols: month, net_sales, total_hours, total_cost, splh, labor_pct
    pricing_trends: SheetPayload  # for KPI: count of items with |pct_change| > 25%
    mtl: SheetPayload             # for KPI: total Shamrock purchase $
    beo: SheetPayload             # for KPI: count of distinct BEO events


_TITLE_FONT = Font(bold=True, size=16, color="1B4332")
_SECTION_FONT = Font(bold=True, size=12, color="2D6A4F")
_KPI_LABEL_FONT = Font(bold=True, size=10, color="555555")
_KPI_VALUE_FONT = Font(bold=True, size=14, color="1B4332")
_KPI_FILL = PatternFill("solid", fgColor="D8F3DC")


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "").replace("%", "")
    try:
        return float(s)
    except ValueError:
        return None


def _yearly_kpis(yearly: SheetPayload, target_year: int) -> dict[str, Any]:
    """Pluck a single year's row out of Toast Yearly Sales."""
    out: dict[str, Any] = {
        "net_sales": None, "labor_cost": None, "labor_pct": None,
        "orders": None, "guests": None,
    }
    if not yearly.rows:
        return out
    # Schema: year, net_sales, gross_sales, orders, checks, guests, ...
    for r in yearly.rows:
        if not r:
            continue
        y = _to_float(r[0])
        if y is None or int(y) != target_year:
            continue
        out["net_sales"] = _to_float(r[1])
        out["orders"] = _to_float(r[3]) if len(r) > 3 else None
        out["guests"] = _to_float(r[5]) if len(r) > 5 else None
        break
    return out


def _labor_year_totals(labor_monthly: SheetPayload) -> dict[str, float | None]:
    """Sum labor cost + net sales across the months in `labor_monthly`. The
    sheet only ever contains the most recent rolling year, so summing it
    gives a year total — no year filter needed."""
    total_cost = 0.0
    total_net = 0.0
    any_row = False
    # Schema: month, net_sales, total_hours, total_cost, splh, labor_pct
    for r in labor_monthly.rows:
        if not r or len(r) < 4:
            continue
        net = _to_float(r[1])
        cost = _to_float(r[3])
        if net is not None:
            total_net += net
            any_row = True
        if cost is not None:
            total_cost += cost
            any_row = True
    if not any_row:
        return {"labor_cost": None, "net_sales": None, "labor_pct": None}
    pct = (total_cost / total_net) if total_net else None
    return {"labor_cost": total_cost, "net_sales": total_net, "labor_pct": pct}


def _shamrock_total_purchases(mtl: SheetPayload) -> float:
    # MTL schema: txn_id, date, type, source_counterparty, ..., total_amount
    total = 0.0
    for r in mtl.rows:
        if not r or len(r) < 10:
            continue
        if r[2] != "Purchase":
            continue
        amt = _to_float(r[9])
        if amt is not None:
            total += amt
    return total


def _beo_event_count(beo: SheetPayload) -> int:
    seen: set[str] = set()
    for r in beo.rows:
        if not r:
            continue
        ef = r[0]
        if ef:
            seen.add(str(ef))
    return len(seen)


def _price_mover_count(pricing: SheetPayload, threshold: float = 0.25) -> int:
    """Count items whose |pct_change| exceeds threshold (default 25%)."""
    n = 0
    # Schema: product_no, description, ..., pct_change (last col)
    for r in pricing.rows:
        if not r:
            continue
        pct = _to_float(r[-1])
        if pct is not None and abs(pct) > threshold:
            n += 1
    return n


def _yearly_sales_chart_block(yearly: SheetPayload) -> list[tuple[int, float]]:
    """Returns (year, net_sales) ascending by year — chart-ready."""
    pairs: list[tuple[int, float]] = []
    for r in yearly.rows:
        if not r:
            continue
        y = _to_float(r[0])
        ns = _to_float(r[1])
        if y is None or ns is None:
            continue
        pairs.append((int(y), ns))
    pairs.sort(key=lambda x: x[0])
    return pairs


def _top_categories_block(
    sales_by_cat: SheetPayload, target_year: int, top_n: int = 10
) -> list[tuple[str, float]]:
    rows: list[tuple[str, float]] = []
    for r in sales_by_cat.rows:
        if not r or len(r) < 3:
            continue
        y = _to_float(r[0])
        if y is None or int(y) != target_year:
            continue
        cat = str(r[1]).strip() if r[1] else None
        ns = _to_float(r[2])
        if not cat or ns is None:
            continue
        rows.append((cat, ns))
    rows.sort(key=lambda x: x[1], reverse=True)
    return rows[:top_n]


def _labor_monthly_block(labor_monthly: SheetPayload) -> list[tuple[int, float]]:
    """Returns (month, labor_pct) ascending by month — chart-ready."""
    pairs: list[tuple[int, float]] = []
    for r in labor_monthly.rows:
        if not r or len(r) < 6:
            continue
        m = _to_float(r[0])
        pct = _to_float(r[5])  # may be a "%" string
        if m is None or pct is None:
            continue
        # Normalize: source emits 27.6 as "27.6%" → /100. If value is already
        # 0.276 we leave it.
        if pct > 1:
            pct = pct / 100
        pairs.append((int(m), pct))
    pairs.sort(key=lambda x: x[0])
    return pairs


def _write_kpi_card(
    ws, *, row: int, col: int, label: str, value: str
) -> None:
    """A KPI card spans two cells: label cell + value cell."""
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = _KPI_LABEL_FONT
    label_cell.fill = _KPI_FILL
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    value_cell = ws.cell(row=row, column=col + 1, value=value)
    value_cell.font = _KPI_VALUE_FONT
    value_cell.fill = _KPI_FILL
    value_cell.alignment = Alignment(horizontal="right", vertical="center")


def _fmt_money(n: float | None) -> str:
    if n is None:
        return "—"
    return f"${n:,.0f}"


def _fmt_int(n: float | None) -> str:
    if n is None:
        return "—"
    return f"{int(round(n)):,}"


def _fmt_pct(p: float | None) -> str:
    if p is None:
        return "—"
    return f"{p * 100:.1f}%"


def _build_dashboard_sheet(wb: Workbook, data: DashboardData) -> None:
    """Insert a 📊 Dashboard sheet at index 1 (after Schema Documentation).

    Layout:
      A1:H1     title
      A4-H7     KPI grid (4 cards × 2 rows)
      A11+      Yearly Net Sales chart + data block (data hidden in cols K-L)
      A28+      Top Categories chart + data block (cols K-L below)
      A45+      Labor % Monthly chart + data block (cols K-L below)
    """
    target_year = 2025  # the most recent full year in the Toast exports

    # Insert at index 1 — Schema Documentation is index 0.
    ws = wb.create_sheet(title="📊 Dashboard", index=1)

    # ── Title ────────────────────────────────────────────────────
    ws.cell(row=1, column=1, value="📊 LARIAT DASHBOARD").font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(
        row=2, column=1,
        value=f"Key metrics from Toast Analytics ({target_year}) + master purchase / event tables.",
    ).font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # ── KPIs ─────────────────────────────────────────────────────
    yearly_kpi = _yearly_kpis(data.yearly_sales, target_year)
    labor_totals = _labor_year_totals(data.labor_monthly)

    ws.cell(row=4, column=1, value="KEY PERFORMANCE INDICATORS").font = _SECTION_FONT
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)

    # Row 5 — top KPI strip (Net Sales / Labor / Labor% / Orders)
    _write_kpi_card(ws, row=5, col=1, label=f"Net Sales {target_year}",
                    value=_fmt_money(yearly_kpi["net_sales"]))
    _write_kpi_card(ws, row=5, col=3, label="Labor Cost (rolling year)",
                    value=_fmt_money(labor_totals["labor_cost"]))
    _write_kpi_card(ws, row=5, col=5, label="Labor % of Net (rolling)",
                    value=_fmt_pct(labor_totals["labor_pct"]))
    _write_kpi_card(ws, row=5, col=7, label=f"Orders {target_year}",
                    value=_fmt_int(yearly_kpi["orders"]))

    # Row 7 — bottom KPI strip (Guests / Shamrock $ / BEO Events / Price Movers)
    _write_kpi_card(ws, row=7, col=1, label=f"Guests {target_year}",
                    value=_fmt_int(yearly_kpi["guests"]))
    _write_kpi_card(ws, row=7, col=3, label="Shamrock Purchase $ (lifetime)",
                    value=_fmt_money(_shamrock_total_purchases(data.mtl)))
    _write_kpi_card(ws, row=7, col=5, label="BEO Events Tracked",
                    value=str(_beo_event_count(data.beo)))
    _write_kpi_card(ws, row=7, col=7, label="Price Movers (>25% Δ)",
                    value=str(_price_mover_count(data.pricing_trends)))

    # KPI cards: widen the value columns so $ figures don't truncate.
    for col in (1, 3, 5, 7):
        ws.column_dimensions[get_column_letter(col)].width = 26
        ws.column_dimensions[get_column_letter(col + 1)].width = 14
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[7].height = 22

    # ── Chart 1: Yearly Net Sales (line) ─────────────────────────
    ws.cell(row=10, column=1, value="YEARLY NET SALES").font = _SECTION_FONT
    yearly_block = _yearly_sales_chart_block(data.yearly_sales)
    _write_chart_data(ws, start_row=11, header=("Year", "Net Sales"),
                      rows=[(y, ns) for y, ns in yearly_block],
                      data_col=11)  # K
    if yearly_block:
        chart = LineChart()
        chart.title = f"Net Sales by Year ({yearly_block[0][0]}–{yearly_block[-1][0]})"
        chart.y_axis.title = "Net Sales ($)"
        chart.x_axis.title = "Year"
        chart.height = 9
        chart.width = 18
        # Data block lives at K11 (header) → K12..K(11+n)
        n = len(yearly_block)
        data_ref = Reference(ws, min_col=12, max_col=12, min_row=11, max_row=11 + n)
        cats_ref = Reference(ws, min_col=11, max_col=11, min_row=12, max_row=11 + n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "A11")

    # ── Chart 2: Top 10 Categories (bar) ─────────────────────────
    ws.cell(row=28, column=1, value=f"TOP 10 SALES CATEGORIES — {target_year}").font = _SECTION_FONT
    cat_block = _top_categories_block(data.sales_by_cat, target_year, top_n=10)
    _write_chart_data(ws, start_row=29, header=("Category", "Net Sales"),
                      rows=cat_block, data_col=11)
    if cat_block:
        chart = BarChart()
        chart.type = "bar"  # horizontal — readable for category labels
        chart.title = f"{target_year} Sales by Category (Top 10)"
        chart.y_axis.title = "Category"
        chart.x_axis.title = "Net Sales ($)"
        chart.height = 11
        chart.width = 18
        n = len(cat_block)
        data_ref = Reference(ws, min_col=12, max_col=12, min_row=29, max_row=29 + n)
        cats_ref = Reference(ws, min_col=11, max_col=11, min_row=30, max_row=29 + n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.legend = None
        ws.add_chart(chart, "A29")

    # ── Chart 3: Labor % Monthly (line) ──────────────────────────
    ws.cell(row=49, column=1, value="LABOR % OF NET SALES — BY MONTH").font = _SECTION_FONT
    labor_block = _labor_monthly_block(data.labor_monthly)
    _write_chart_data(ws, start_row=50, header=("Month", "Labor % of Net"),
                      rows=labor_block, data_col=11)
    if labor_block:
        chart = LineChart()
        chart.title = "Labor % of Net Sales by Month (rolling year)"
        chart.y_axis.title = "Labor % of Net Sales"
        chart.x_axis.title = "Month (1–12)"
        chart.height = 9
        chart.width = 18
        n = len(labor_block)
        data_ref = Reference(ws, min_col=12, max_col=12, min_row=50, max_row=50 + n)
        cats_ref = Reference(ws, min_col=11, max_col=11, min_row=51, max_row=50 + n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "A50")

    # Hide the data-block column band (K=11, L=12) so the dashboard reads
    # as cards + charts; the data is still present for the charts to bind to.
    ws.column_dimensions["K"].hidden = True
    ws.column_dimensions["L"].hidden = True
    # Pin top rows so KPIs are always visible while scrolling.
    ws.freeze_panes = "A4"


def _write_chart_data(
    ws, *, start_row: int, header: tuple[str, str],
    rows: list[tuple[Any, Any]], data_col: int,
) -> None:
    """Write a small (label, value) data block with header into two
    consecutive columns starting at (start_row, data_col)."""
    ws.cell(row=start_row, column=data_col, value=header[0]).font = Font(bold=True)
    ws.cell(row=start_row, column=data_col + 1, value=header[1]).font = Font(bold=True)
    for i, r in enumerate(rows, start=1):
        ws.cell(row=start_row + i, column=data_col, value=r[0])
        ws.cell(row=start_row + i, column=data_col + 1, value=r[1])


# ── Workbook writer ──────────────────────────────────────────────


def write_workbook(
    path: Path,
    sheets: Iterable[SheetPayload],
    *,
    dashboard: "DashboardData | None" = None,
) -> None:
    wb = Workbook()
    # Remove the default sheet; we'll add our own.
    default = wb.active
    wb.remove(default)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")

    for s in sheets:
        ws = wb.create_sheet(title=_truncate_sheet_name(s.name))
        ws.append(s.header)
        for col_idx in range(1, len(s.header) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for row in s.rows:
            ws.append(row)
        # Reasonable column widths.
        for i, col_name in enumerate(s.header, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(
                12, min(48, len(str(col_name)) + 4)
            )
        if s.freeze_top:
            ws.freeze_panes = "A2"

    # Dashboard sheet is built last so we can read SheetPayloads, but moved
    # to position 1 (right after Schema Documentation) for visibility.
    if dashboard is not None:
        _build_dashboard_sheet(wb, dashboard)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def _truncate_sheet_name(name: str) -> str:
    """Excel hard-limits sheet names to 31 chars."""
    if len(name) <= 31:
        return name
    return name[:28] + "…"


# ── Main ─────────────────────────────────────────────────────────


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--archive", default=DEFAULT_ARCHIVE)
    p.add_argument("--output", default=DEFAULT_OUTPUT)
    args = p.parse_args(argv)

    archive = Path(args.archive)
    output = Path(args.output)
    master_xlsx = archive / "XL" / "Lariat Master Workbook.xlsx"

    if not archive.exists():
        print(f"archive not found: {archive}", file=sys.stderr)
        return 2
    if not master_xlsx.exists():
        print(f"master workbook not found: {master_xlsx}", file=sys.stderr)
        return 2

    print(f"[build] archive={archive}")
    print(f"[build] master={master_xlsx}")
    print(f"[build] output={output}")

    # Parse reference sources first — downstream derived sheets read from these.
    shamrock = build_shamrock_price_list_ref(master_xlsx)
    print(f"  shamrock PL rows: {shamrock.row_count()}")
    sysco_ph = build_sysco_purchase_history_ref(archive)
    print(f"  sysco PH rows:    {sysco_ph.row_count()}")
    sysco_beo = build_sysco_beo_basics_ref(archive)
    print(f"  sysco BEO rows:   {sysco_beo.row_count()}")
    toast_menu = build_toast_menu_items_ref(archive)
    print(f"  toast menu rows:  {toast_menu.row_count()}")
    ws_spend = build_webstaurantstore_ref(archive)
    print(f"  webstaur. rows:   {ws_spend.row_count()}")
    beo = build_beo_invoices(archive)
    print(f"  BEO invoice rows: {beo.row_count()}")

    toast_analytics = build_toast_analytics(archive)
    for s in toast_analytics:
        print(f"  {s.name}: {s.row_count()} rows")

    shamrock_orders = build_shamrock_orders(master_xlsx)
    print(f"  shamrock orders:        {shamrock_orders.row_count()} rows")
    beo_prep = build_beo_prep_ref(master_xlsx)
    print(f"  BEO prep rows:          {beo_prep.row_count()}")

    # Derived sheets.
    mpc = build_master_product_catalog(shamrock, sysco_ph, sysco_beo)
    print(f"  master product catalog: {mpc.row_count()} rows")
    comparison = build_supplier_comparison(shamrock, sysco_ph)
    print(f"  supplier comparison:    {comparison.row_count()} rows")
    mtl = build_master_transaction_log(beo, shamrock_orders)
    print(f"  master transaction log: {mtl.row_count()} rows")
    
    pricing_trends = build_pricing_trends(shamrock_orders)
    print(f"  pricing trends:         {pricing_trends.row_count()} rows")

    deferred = build_deferred_todos()

    # Assemble in display order.
    ordered: list[SheetPayload] = [
        # Schema docs first — but we need row counts, so build after everything.
        mpc,
        mtl,
        shamrock_orders,
        pricing_trends,
        beo,
        beo_prep,
        comparison,
        *toast_analytics,
        shamrock,
        sysco_ph,
        sysco_beo,
        toast_menu,
        ws_spend,
        deferred,
    ]
    counts = {s.name: s.row_count() for s in ordered}
    schema = build_schema_documentation(counts)
    ordered.insert(0, schema)

    # Dashboard inputs — look up by sheet name so the dashboard builder
    # never has to know which CSV produced which payload.
    by_name = {s.name: s for s in toast_analytics}
    dashboard = DashboardData(
        yearly_sales=by_name.get(
            "📊 Toast Analytics — Yearly Sales",
            SheetPayload("📊 Toast Analytics — Yearly Sales", []),
        ),
        sales_by_cat=by_name.get(
            "📊 Toast Analytics — Sales By Category",
            SheetPayload("📊 Toast Analytics — Sales By Category", []),
        ),
        labor_monthly=by_name.get(
            "📊 Toast Analytics — Labor Monthly",
            SheetPayload("📊 Toast Analytics — Labor Monthly", []),
        ),
        pricing_trends=pricing_trends,
        mtl=mtl,
        beo=beo,
    )

    write_workbook(output, ordered, dashboard=dashboard)
    print(f"\n✅ wrote {output}  ({len(ordered) + 1} sheets, +1 dashboard)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
