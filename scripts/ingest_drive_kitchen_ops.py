#!/usr/bin/env python3
"""Ingest the official kitchen-ops files (Drive folder) into Lariat caches + DB.

Reads xlsx / docx files staged under ``data/imports/drive-kitchen-ops-20260421/``
and writes to:

- ``data/cache/line_checks.json`` — adds ``expo`` key (items from Expo Line Check.xlsx).
- ``data/cache/stations.json`` — sets the expo station's ``line_check_key`` to 'expo'.
- ``data/cache/closings.json`` — new cache. Per-station and house closing checklists.
- ``data/cache/weekly_prep.json`` — new cache. By-day and by-category prep lists.
- ``data/cache/order_guide.json`` — new cache. Sysco order guide items with PAR.
- ``cleaning_schedule`` table — inserts Weekly / Monthly / Daily (dish pit) tasks.
  Idempotent: dedupes on ``(location_id, area, task, frequency, active)``.

Intentionally skipped:
- ``sean-admin_credentials.csv`` (credentials; never read).
- ``Brunch Line (working).xlsx``, ``Diedre 7_22 Kitchen.xlsx`` (single-day instances,
  not templates).
- ``FRYER_LOWBOY TOP STATION.xlsx``, ``FRYRER_FREEZER DIAGRAM.xlsx`` (station
  labels; add as notes by hand if useful).
- ``LINE CHECK FRY_.xlsx``, ``LINE CHECK GARDE.xlsx`` (older, superseded by the
  longer par/have/need versions already in line_checks.json).
- ``CLOSING GRILL.xlsx`` / ``CLOSING SALAD.xlsx`` (superseded by the per-station
  closings in ``Closing checklist_ Positions_.docx``, which covers all 4 stations).
- ``GRILL SET-UP.xlsx`` / ``SALAD SET-UP.xlsx`` / ``BRUNCH SET UP.xlsx`` and
  ``Setups.docx`` (setups.json already derived from the canonical Setups.docx;
  rerun ingest-costing to refresh if Setups.docx changes).
- ``Coleslaw Recipe 8_13.docx`` — matches the existing recipe in recipes.json 1:1.
- ``WEEKLY PREP.xlsx`` — the .docx is the canonical prep list (items + grouping
  by day). The .xlsx is the empty par/stock/need form used at the board.

Usage::

    python3 scripts/ingest_drive_kitchen_ops.py          # write
    python3 scripts/ingest_drive_kitchen_ops.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DB = ROOT / "data" / "lariat.db"
CACHE = ROOT / "data" / "cache"
IMPORTS = ROOT / "data" / "imports" / "drive-kitchen-ops-20260421"

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
TIME_LINE_RE = re.compile(r"^\s*\d{1,2}[:.]\d{2}(?:\s*[-–]\s*\d{1,2}[:.]\d{2})?\s*$")


# ── readers ───────────────────────────────────────────────────────

def docx_paragraphs(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as z:
        with z.open("word/document.xml") as f:
            tree = ET.parse(f)
    out: list[str] = []
    for p in tree.iter(W + "p"):
        txt = "".join(t.text or "" for t in p.iter(W + "t")).strip()
        if txt:
            out.append(txt)
    return out


def xlsx_col0_nonblank(path: Path) -> list[str]:
    """Return non-blank first-column strings, excluding obvious time markers."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    out: list[str] = []
    for r in ws.iter_rows(values_only=True):
        v = r[0]
        if v is None:
            continue
        s = str(v).strip()
        if not s or TIME_LINE_RE.match(s):
            continue
        out.append(s)
    return out


# ── extractors ────────────────────────────────────────────────────

def extract_expo_items() -> list[str]:
    """Expo Line Check.xlsx → item list, dropping header + blanks."""
    rows = xlsx_col0_nonblank(IMPORTS / "Expo Line Check.xlsx")
    # Drop banner + header row labels.
    return [r for r in rows if r.lower() not in {"expo line check", "item"}]


def extract_closings() -> dict[str, list[str]]:
    """Parse Closing checklist_ Positions_.docx into per-station sections.
    Section headers look like ``Closing checklist: Garde`` / ``Closing list: Expo``.
    Also captures the general Closing procedures.docx as the ``house`` key.
    """
    lines = docx_paragraphs(IMPORTS / "Closing checklist_ Positions_.docx")
    out: dict[str, list[str]] = {}
    current_key: str | None = None
    header_re = re.compile(r"^Closing (?:checklist|list)\s*:\s*(.+?)\s*$", re.I)
    key_map = {
        "garde": "garde",
        "fry": "fry",
        "grill/sautee": "grill_saute",
        "grill/saute": "grill_saute",
        "grille/saute": "grill_saute",
        "expo": "expo",
    }
    for ln in lines:
        m = header_re.match(ln)
        if m:
            raw = m.group(1).strip().lower()
            current_key = key_map.get(raw, raw.replace(" ", "_").replace("/", "_"))
            out.setdefault(current_key, [])
            continue
        if current_key:
            out[current_key].append(ln)

    # House close-out (whole-kitchen).
    house_raw = docx_paragraphs(IMPORTS / "Closing procedures.docx")
    house = []
    skipped_header = False
    for ln in house_raw:
        if not skipped_header and ln.lower().startswith("closing procedures"):
            skipped_header = True
            continue
        # Strip leading "1." / "1)" numbering.
        cleaned = re.sub(r"^\s*\d+\s*[.)]\s*", "", ln).strip()
        if cleaned:
            house.append(cleaned)
    # The last item is a morale joke; drop anything that doesn't read as a task.
    house = [h for h in house if not re.search(r"get bitches", h, re.I)]
    out["house"] = house
    return out


def extract_weekly_prep() -> dict[str, object]:
    """Weekly Prep.docx = day-grouped list. Prep list.docx = category list."""
    day_lines = docx_paragraphs(IMPORTS / "Weekly Prep.docx")
    by_day: dict[str, list[str]] = {}
    cur_day: str | None = None
    day_header_re = re.compile(r"^Weekly prep\s+(\w+)\s*$", re.I)
    for ln in day_lines:
        if ln.lower().startswith("tab "):
            continue
        m = day_header_re.match(ln)
        if m:
            cur_day = m.group(1).capitalize()
            by_day.setdefault(cur_day, [])
            continue
        if cur_day:
            by_day[cur_day].append(ln)

    cat_lines = docx_paragraphs(IMPORTS / "Prep list.docx")
    by_category: dict[str, list[str]] = {}
    cur_cat: str | None = None
    known_cats = {"Sauces", "Dressings", "Butters", "Starches", "Veg Prep",
                  "Pickles and Spice", "Proteins", "Breakfast", "Dairy", "Dessert"}
    for ln in cat_lines:
        if ln.lower() == "prep list":
            continue
        cleaned = re.sub(r"\s*_+\s*$", "", ln).strip()
        if cleaned in known_cats:
            cur_cat = cleaned
            by_category.setdefault(cur_cat, [])
            continue
        if cur_cat:
            by_category[cur_cat].append(cleaned)

    return {"by_day": by_day, "by_category": by_category}


def extract_order_guide() -> list[dict]:
    """Shop_Order guide summer 25.xlsx — Sysco SUPC rows with PAR levels."""
    path = IMPORTS / "Shop_Order guide summer 25_059_075356.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c else "" for c in rows[0]]
    # Expected: SUPC, Desc, Size, Brand, Unit, Cat, Location, On Hand, Par, Need
    items = []
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        rec = {h: r[i] for i, h in enumerate(header) if i < len(r) and h}
        supc = rec.get("SUPC")
        desc = rec.get("Desc")
        if supc is None or not desc:
            continue
        if isinstance(supc, float) and supc.is_integer():
            supc_str = str(int(supc))
        else:
            supc_str = str(supc).strip()
        items.append({
            "supc": supc_str,
            "description": str(desc).strip(),
            "pack_size": str(rec.get("Size") or "").strip() or None,
            "brand": str(rec.get("Brand") or "").strip() or None,
            "unit": str(rec.get("Unit") or "").strip() or None,
            "category": str(rec.get("Cat") or "").strip() or None,
            "location": str(rec.get("Location") or "").strip() or None,
            "par": str(rec.get("Par") or "").strip() or None,
        })
    return items


def extract_cleaning_schedule_rows() -> list[dict]:
    """Build cleaning_schedule rows from Monthly / Weekly / Dishwasher sources."""
    rows: list[dict] = []

    # Monthly
    monthly = docx_paragraphs(IMPORTS / "Monthly Cleaning_Maintenance.docx")
    for ln in monthly:
        if ln.lower().startswith("monthly cleaning"):
            continue
        rows.append({
            "area": "General",
            "task": ln,
            "frequency": "Monthly",
            "notes": None,
        })

    # Weekly (grouped by day-of-week header)
    weekly = docx_paragraphs(IMPORTS / "Weekly Cleaning.docx")
    days = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
            "Saturday", "Sunday"}
    cur_day: str | None = None
    for ln in weekly:
        if ln.lower() == "weekly cleaning":
            continue
        if ln.strip() in days:
            cur_day = ln.strip()
            continue
        rows.append({
            "area": "General",
            "task": ln,
            "frequency": "Weekly",
            "notes": f"Standing {cur_day} task." if cur_day else None,
        })

    # Dishwasher — daily, time-stamped. Drop the time headers.
    dish = xlsx_col0_nonblank(IMPORTS / "DISHWASHER DUTIES.xlsx")
    for task in dish:
        rows.append({
            "area": "Dish Pit",
            "task": task,
            "frequency": "Daily",
            "notes": "Dishwasher duty.",
        })

    return rows


# ── writers ───────────────────────────────────────────────────────

def update_line_checks(dry: bool) -> tuple[int, str]:
    p = CACHE / "line_checks.json"
    data = json.loads(p.read_text())
    expo = extract_expo_items()
    before = data.get("expo", [])
    if before == expo:
        return 0, f"expo line check unchanged ({len(expo)} items)"
    data["expo"] = expo
    if not dry:
        p.write_text(json.dumps(data, indent=2))
    return len(expo), f"expo line check → {len(expo)} items (was {len(before)})"


def update_stations(dry: bool) -> str:
    p = CACHE / "stations.json"
    data = json.loads(p.read_text())
    changed = False
    for s in data:
        if s.get("id") == "expo" and s.get("line_check_key") != "expo":
            s["line_check_key"] = "expo"
            changed = True
    if changed and not dry:
        p.write_text(json.dumps(data, indent=2))
    return ("stations.expo.line_check_key → 'expo'" if changed
            else "stations unchanged")


def write_closings(dry: bool) -> str:
    data = extract_closings()
    p = CACHE / "closings.json"
    if not dry:
        p.write_text(json.dumps(data, indent=2))
    summary = ", ".join(f"{k}={len(v)}" for k, v in data.items())
    return f"closings.json → {summary}"


def write_weekly_prep(dry: bool) -> str:
    data = extract_weekly_prep()
    p = CACHE / "weekly_prep.json"
    if not dry:
        p.write_text(json.dumps(data, indent=2))
    days = ", ".join(f"{k}:{len(v)}" for k, v in data["by_day"].items())
    cats = ", ".join(f"{k}:{len(v)}" for k, v in data["by_category"].items())
    return f"weekly_prep.json → by_day({days}) by_category({cats})"


def write_order_guide(dry: bool) -> str:
    items = extract_order_guide()
    p = CACHE / "order_guide.json"
    if not dry:
        p.write_text(json.dumps({"items": items}, indent=2))
    return f"order_guide.json → {len(items)} SUPC rows"


def seed_cleaning_schedule(db_path: Path, dry: bool) -> str:
    rows = extract_cleaning_schedule_rows()
    if dry:
        return f"cleaning_schedule would seed {len(rows)} rows (not written)"
    con = sqlite3.connect(db_path)
    try:
        inserted = 0
        skipped = 0
        for r in rows:
            exists = con.execute(
                """SELECT id FROM cleaning_schedule
                    WHERE location_id = 'default'
                      AND area = ? AND task = ? AND frequency = ?
                      AND active = 1""",
                (r["area"], r["task"], r["frequency"]),
            ).fetchone()
            if exists:
                skipped += 1
                continue
            con.execute(
                """INSERT INTO cleaning_schedule
                     (location_id, area, task, frequency, notes, active)
                   VALUES ('default', ?, ?, ?, ?, 1)""",
                (r["area"], r["task"], r["frequency"], r["notes"]),
            )
            inserted += 1
        con.commit()
    finally:
        con.close()
    return f"cleaning_schedule → inserted {inserted}, skipped-existing {skipped}"


# ── main ──────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--db", type=Path, default=DEFAULT_DB)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not IMPORTS.exists():
        print(f"Missing source dir: {IMPORTS}", file=sys.stderr)
        return 2

    print(f"Reading from {IMPORTS.relative_to(ROOT)}  "
          f"({'dry-run' if args.dry_run else 'writing'})")

    summaries: list[str] = []
    n_expo, s = update_line_checks(args.dry_run); summaries.append(s)
    summaries.append(update_stations(args.dry_run))
    summaries.append(write_closings(args.dry_run))
    summaries.append(write_weekly_prep(args.dry_run))
    summaries.append(write_order_guide(args.dry_run))
    summaries.append(seed_cleaning_schedule(args.db, args.dry_run))

    print("\nKitchen-ops ingest summary:")
    for s in summaries:
        print(f"  - {s}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
