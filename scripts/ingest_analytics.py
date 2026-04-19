#!/usr/bin/env python3
"""Parse unified workbook Toast sales + analytics spend; print JSON."""
import json
import os
import re
import sys

import openpyxl

UNIFIED = os.environ.get("LARIAT_UNIFIED", "")
ANALYTICS = os.environ.get("LARIAT_ANALYTICS", "")

out = {"sales_lines": [], "spend_monthly": [], "toast_sheet": None}

# Pick first matching Toast - Item Sales sheet
def find_item_sales_sheet(wb):
    for name in wb.sheetnames:
        if name.startswith("Toast - Item Sales"):
            return name
    return None


def load_unified(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sn = find_item_sales_sheet(wb)
    if not sn:
        return
    out["toast_sheet"] = sn
    ws = wb[sn]
    started = False
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        a = str(row[0]).strip()
        if a.lower() == "item name":
            started = True
            continue
        if not started:
            continue
        qty, rev = row[1], row[2]
        out["sales_lines"].append(
            {
                "item_name": a,
                "quantity_sold": float(qty) if isinstance(qty, (int, float)) else qty,
                "net_sales": float(rev) if isinstance(rev, (int, float)) else rev,
            }
        )


def load_analytics_spend(path):
    if not os.path.exists(path):
        return
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    name = "📈 Monthly Spend"
    if name not in wb.sheetnames:
        return
    ws = wb[name]
    started = False
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        # columns B-E: Month, Total Spend, ...
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        if b is not None and str(b).strip().lower() == "month":
            started = True
            continue
        if not started:
            continue
        if b is None or c is None:
            continue
        month = str(b).strip()
        if not re.match(r"^\d{4}-\d{2}$", month):
            continue
        amt = float(c) if isinstance(c, (int, float)) else c
        out["spend_monthly"].append({"month": month, "shamrock_total_spend": amt, "source": "analytics_workbook"})


if UNIFIED and os.path.exists(UNIFIED):
    load_unified(UNIFIED)
if ANALYTICS and os.path.exists(ANALYTICS):
    load_analytics_spend(ANALYTICS)

print(json.dumps(out))
sys.stdout.flush()
