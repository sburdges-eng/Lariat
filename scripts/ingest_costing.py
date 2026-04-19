#!/usr/bin/env python3
"""Parse Master Costing + operations Order Guide; print JSON for ingest-costing.mjs."""
import json
import os
import re
import sys

import openpyxl

COSTING = os.environ.get("LARIAT_COSTING", "")
OPS = os.environ.get("LARIAT_OPS", "")

out = {
    "vendor_prices": [],
    "recipe_costs": [],
    "bom_lines": [],
    "ingredient_maps": [],
    "order_guide": [],
}


def load_costing(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Vendor Prices" in wb.sheetnames:
        ws = wb["Vendor Prices"]
        started = False
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            a = str(row[0]).strip()
            if a.lower() == "ingredient":
                started = True
                continue
            if not started:
                continue
            if len(row) < 7:
                continue
            ing, vend, sku, psize, pun, pprice, uprice = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
            cat = row[7] if len(row) > 7 else None
            out["vendor_prices"].append(
                {
                    "ingredient": str(ing).strip() if ing else "",
                    "vendor": str(vend).strip().lower() if vend else "",
                    "sku": str(sku) if sku is not None else "",
                    "pack_size": float(psize) if isinstance(psize, (int, float)) else psize,
                    "pack_unit": str(pun).strip() if pun else "",
                    "pack_price": float(pprice) if isinstance(pprice, (int, float)) else pprice,
                    "unit_price": float(uprice) if isinstance(uprice, (int, float)) else uprice,
                    "category": str(cat).strip() if cat else None,
                }
            )

    if "Recipe Cost Summary" in wb.sheetnames:
        ws = wb["Recipe Cost Summary"]
        started = False
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            if str(row[0]).strip().lower() == "recipe id":
                started = True
                continue
            if not started:
                continue
            rid, rname, cat, yld, yunit = row[0], row[1], row[2], row[3], row[4]
            if rid is None or str(rid).strip() == "" or str(rid).strip().startswith("▸"):
                continue
            bcost, cyu, cl, tl, interp = row[5], row[6], row[7], row[8], row[9] if len(row) > 9 else (None,) * 5
            out["recipe_costs"].append(
                {
                    "recipe_id": str(rid).strip() if rid else "",
                    "recipe_name": str(rname).strip() if rname else "",
                    "category": str(cat).strip() if cat else "",
                    "yield": float(yld) if isinstance(yld, (int, float)) else yld,
                    "yield_unit": str(yunit).strip() if yunit else "",
                    "batch_cost": float(bcost) if isinstance(bcost, (int, float)) else bcost,
                    "cost_per_yield_unit": float(cyu) if isinstance(cyu, (int, float)) else cyu,
                    "costed_lines": int(cl) if isinstance(cl, (int, float)) else cl,
                    "total_lines": int(tl) if isinstance(tl, (int, float)) else tl,
                    "interpretations": int(interp) if isinstance(interp, (int, float)) else interp,
                }
            )

    if "Full BOM Detail" in wb.sheetnames:
        ws = wb["Full BOM Detail"]
        started = False
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            a0 = row[0]
            if a0 is None:
                continue
            s0 = str(a0).strip()
            if s0.lower() == "recipe id":
                started = True
                continue
            if not started:
                continue
            if s0.startswith("▸") or s0.startswith("LARIAT"):
                continue
            rid, ing, qty, unit = row[0], row[1], row[2], row[3]
            subr, ving, mstat, vend = (
                (row[4], row[5], row[6], row[7]) if len(row) > 7 else (None, None, None, None)
            )
            pp, psz = (row[8], row[9]) if len(row) > 9 else (None, None)
            out["bom_lines"].append(
                {
                    "recipe_id": str(rid).strip() if rid else "",
                    "ingredient": str(ing).strip() if ing else "",
                    "qty": float(qty) if isinstance(qty, (int, float)) else qty,
                    "unit": str(unit).strip() if unit else "",
                    "sub_recipe": str(subr).strip() if subr else None,
                    "vendor_ingredient": str(ving).strip() if ving else None,
                    "map_status": str(mstat).strip() if mstat else None,
                    "vendor": str(vend).strip().lower() if vend else None,
                    "pack_price": float(pp) if isinstance(pp, (int, float)) else pp,
                    "pack_size": float(psz) if isinstance(psz, (int, float)) else psz,
                }
            )

    if "Ingredient Vendor Map" in wb.sheetnames:
        ws = wb["Ingredient Vendor Map"]
        started = False
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            if str(row[0]).strip().lower() == "recipe ingredient":
                started = True
                continue
            if not started:
                continue
            out["ingredient_maps"].append(
                {
                    "recipe_ingredient": str(row[0]).strip(),
                    "vendor_ingredient": str(row[1]).strip() if row[1] else "",
                    "status": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                }
            )


def load_order_guide(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Order Guide" not in wb.sheetnames:
        return
    ws = wb["Order Guide"]
    started = False
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        if str(row[0]).strip().lower() == "ingredient":
            started = True
            continue
        if not started:
            continue
        ing, bq, u, sq, vend, up, ext = row[0], row[1], row[2], row[3], row[4], row[5], row[6] if len(row) > 6 else None
        out["order_guide"].append(
            {
                "ingredient": str(ing).strip(),
                "base_qty": float(bq) if isinstance(bq, (int, float)) else bq,
                "unit": str(u).strip() if u else "",
                "vendor": str(vend).strip().lower() if vend else "",
                "unit_price": float(up) if isinstance(up, (int, float)) else up,
            }
        )


if COSTING and os.path.exists(COSTING):
    load_costing(COSTING)
if OPS and os.path.exists(OPS):
    load_order_guide(OPS)

print(json.dumps(out))
sys.stdout.flush()
