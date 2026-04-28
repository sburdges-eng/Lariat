"""Build a deterministic minimal xlsx fixture for ingest_shows_xlsx tests.

Sheets:
- future: 5 rows (1 duplicate band, 1 with 'w' status, 1 with 'jb, bit, sk')
- past:   8 rows (year banners 2025 + 2024, 1 malformed missing date)
- tiktok plan: 3 structured + 1 idea-only-note row

The xlsx itself is gitignored; this script regenerates it on demand.
Run: python3 tests/python/fixtures/build_shows_fixture.py
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path

from openpyxl import Workbook

OUT = Path(__file__).parent / "shows_minimal.xlsx"


def build() -> Path:
    wb = Workbook()
    ws_future = wb.active
    ws_future.title = "future"

    # Header row matches the real workbook headers exactly.
    ws_future.append([
        "Band Name", "Date", "Media list", "MKTing adv", "Auto-counts",
        "Announce date", "Meta ads", "FB event", "Co-host sent",
        "create DICE tickets", "Listing on Jambase, BIT, Songkick",
        "DICE email (ticket sale, DOS)", "Newsletter (weekly, monthly)",
        "Assets", "Posts", "WHBV", "Door tix", "Price", None,
    ])

    # 5 data rows — only the trailing 18 cells; col 19 is None.
    ws_future.append([
        "openmic dinnershow", _dt.datetime(2026, 5, 1),
        "-", "-", "-", "-", "-", "y", "-", "y", "-", "-", "-", "y", "-", "n", "-", 0.0, None,
    ])
    ws_future.append([
        "openmic dinnershow", _dt.datetime(2026, 5, 8),  # duplicate band, different date
        "-", "-", "-", "-", "-", "y", "-", "y", "-", "-", "-", "y", "-", "n", "-", 0.0, None,
    ])
    ws_future.append([
        "armchair boogie", _dt.datetime(2026, 5, 15),
        "y", "y", "n", "y", "y", "y", "accepted", "y", "jb, bit, sk",
        "tix, dos", "w", "y", 6.0, "n", "y", 15.0, None,
    ])
    ws_future.append([
        "the bramble hollow", _dt.datetime(2026, 5, 22),
        "-", "y", "-", "y", "y", "y", "pending", "n", "-", "-", "-", "y", 0, "n", "-", 12.0, None,
    ])
    ws_future.append([
        "junior and the aces", _dt.datetime(2026, 6, 1),
        "y", "y", "y", "y", "y", "y", "y", "y", "jb, bit, sk", "tix, dos", "y", "y", 12.0, "y", "y", 18.0, None,
    ])

    # `past` sheet: year-banner row format, sparse columns.
    ws_past = wb.create_sheet("past")
    ws_past.append([2025])  # year banner row
    ws_past.append(["open mic", None, _dt.datetime(2025, 2, 26)])
    ws_past.append(["karaoke", None, _dt.datetime(2025, 2, 27)])
    ws_past.append(["pete n mark", None, _dt.datetime(2025, 2, 28)])
    ws_past.append(["malformed-no-date", None, None])  # dropped row
    ws_past.append([2024])  # year banner row
    ws_past.append(["the hip snacks", None, _dt.datetime(2024, 3, 1)])
    ws_past.append(["the whiskey sweets brunch", None, _dt.datetime(2024, 3, 2)])

    # `tiktok plan` sheet: 3 structured + 1 idea-only-note row.
    ws_tt = wb.create_sheet("tiktok plan")
    ws_tt.append(["idea", "video content", "staff needed", "props etc needed", None])
    ws_tt.append([
        "introducing your new favorite music venue",
        "walking around venue",
        "bartenders, band, crowd",
        "drinks being poured",
        None,
    ])
    ws_tt.append([
        "aesthetic cocktail recipe",
        "closeups behind bar",
        "lauren",
        "the bar",
        None,
    ])
    ws_tt.append([
        "almost forgot that this was the point",
        "clips from shows",
        "na",
        "na",
        "***can make this with existing videos",
    ])
    ws_tt.append([
        "thoughts on a 'lauren at the lariat' tiktok account",
        None, None, None, None,  # idea-only — routes to notes
    ])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    p = build()
    print(p)
