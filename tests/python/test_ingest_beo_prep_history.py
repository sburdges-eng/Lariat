"""Unit tests for scripts.ingest_beo_prep_history.

Builds a tiny in-memory xlsx fixture matching the master workbook's
`📋 BEO Prep` schema, runs the parser + upsert, and asserts:

  - header row is detected and skipped
  - blank rows are skipped
  - Main Item rows preserve numeric Amount/Qty as text
  - Secondary Prep / Special Sauce rows preserve descriptive text
  - rows missing an item are dropped (table NOT NULL)
  - upsert is a full refresh: re-running with a different fixture
    replaces the rows for (location_id, source) only — other source
    labels stay untouched
  - openpyxl datetime in 'Event Date' lands as ISO YYYY-MM-DD
"""
from __future__ import annotations

import datetime as dt
import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.ingest_beo_prep_history import (  # noqa: E402
    SHEET_NAME,
    load_rows,
    upsert,
)

DDL = """
CREATE TABLE IF NOT EXISTS beo_prep_history (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  location_id TEXT NOT NULL DEFAULT 'default',
  client TEXT,
  event_date TEXT,
  event_file TEXT,
  type TEXT,
  item TEXT NOT NULL,
  amount_qty TEXT,
  prep_day TEXT,
  pre_prep_notes TEXT,
  plating_notes TEXT,
  source TEXT NOT NULL,
  imported_at TEXT DEFAULT (datetime('now'))
);
"""


def _make_db(path: Path) -> None:
    con = sqlite3.connect(str(path))
    try:
        con.executescript(DDL)
        con.commit()
    finally:
        con.close()


def _make_fixture_xlsx(
    path: Path, rows: list[list], sheet_name: str = SHEET_NAME
) -> None:
    """Write rows under a header row to ``sheet_name`` in a new xlsx."""
    wb = openpyxl.Workbook()
    # Remove the default sheet, then add ours with the exact emoji name.
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet(sheet_name)
    ws.append(
        [
            "Client",
            "Event Date",
            "Type",
            "Item",
            "Amount/Qty",
            "Prep Day",
            "Pre-Prep",
            "Plating",
        ]
    )
    for r in rows:
        ws.append(r)
    wb.save(str(path))
    wb.close()


class LoadRows(unittest.TestCase):
    def test_main_item_row_preserved(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "fixture.xlsx"
            _make_fixture_xlsx(
                xlsx,
                [
                    [
                        "Darrell & Anne Collett",
                        dt.datetime(2025, 9, 27),
                        "Main Item",
                        "Mac Balls",
                        50,
                        "Saturday",
                        "ditalini cooked fri",
                        "4 inch plate",
                    ],
                ],
            )
            rows = load_rows(xlsx)
            self.assertEqual(len(rows), 1)
            r = rows[0]
            self.assertEqual(r["client"], "Darrell & Anne Collett")
            self.assertEqual(r["event_date"], "2025-09-27")
            self.assertEqual(r["type"], "Main Item")
            self.assertEqual(r["item"], "Mac Balls")
            self.assertEqual(r["amount_qty"], "50")
            self.assertEqual(r["prep_day"], "Saturday")
            self.assertEqual(r["pre_prep_notes"], "ditalini cooked fri")
            self.assertEqual(r["plating_notes"], "4 inch plate")

    def test_descriptive_amount_qty_preserved_as_text(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "fixture.xlsx"
            _make_fixture_xlsx(
                xlsx,
                [
                    [
                        "Darrell & Anne Collett",
                        dt.datetime(2025, 9, 27),
                        "Secondary Prep",
                        "Queso",
                        "Special Sauce",
                        None,
                        None,
                        None,
                    ],
                ],
            )
            rows = load_rows(xlsx)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["amount_qty"], "Special Sauce")

    def test_blank_rows_and_missing_item_skipped(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "fixture.xlsx"
            _make_fixture_xlsx(
                xlsx,
                [
                    [None, None, None, None, None, None, None, None],  # blank
                    [
                        "Anyone",
                        dt.datetime(2025, 9, 27),
                        "Main Item",
                        None,  # missing item — should drop
                        10,
                        None,
                        None,
                        None,
                    ],
                    [
                        "Anyone",
                        dt.datetime(2025, 9, 27),
                        "Main Item",
                        "Real Item",
                        10,
                        None,
                        None,
                        None,
                    ],
                ],
            )
            rows = load_rows(xlsx)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["item"], "Real Item")

    def test_missing_sheet_returns_empty(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "fixture.xlsx"
            # Build a workbook that does NOT contain the BEO Prep sheet.
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Other"
            ws.append(["a", "b", "c"])
            wb.save(str(xlsx))
            wb.close()
            self.assertEqual(load_rows(xlsx), [])

    def test_integer_float_amount_qty_stripped_to_int_text(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "fixture.xlsx"
            _make_fixture_xlsx(
                xlsx,
                [
                    [
                        "X",
                        dt.datetime(2025, 1, 1),
                        "Main Item",
                        "Foo",
                        50.0,  # openpyxl writes ints as floats sometimes
                        None,
                        None,
                        None,
                    ],
                ],
            )
            rows = load_rows(xlsx)
            self.assertEqual(rows[0]["amount_qty"], "50")


class UpsertRefresh(unittest.TestCase):
    def test_full_refresh_replaces_only_matching_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            db = tmpdir / "lariat.db"
            _make_db(db)

            # Pre-seed an unrelated source row that should NOT be touched.
            con = sqlite3.connect(str(db))
            try:
                con.execute(
                    """INSERT INTO beo_prep_history
                         (location_id, item, source) VALUES (?, ?, ?)""",
                    ("default", "Untouched", "other_source"),
                )
                con.commit()
            finally:
                con.close()

            xlsx = tmpdir / "fixture.xlsx"
            _make_fixture_xlsx(
                xlsx,
                [
                    [
                        "X", dt.datetime(2025, 1, 1), "Main Item", "Item A",
                        10, None, None, None,
                    ],
                    [
                        "X", dt.datetime(2025, 1, 1), "Main Item", "Item B",
                        20, None, None, None,
                    ],
                ],
            )
            rows = load_rows(xlsx)
            n = upsert(db, rows, location_id="default", source="run1")
            self.assertEqual(n, 2)

            # Re-run with a smaller fixture under the same source — replaces.
            xlsx2 = tmpdir / "fixture2.xlsx"
            _make_fixture_xlsx(
                xlsx2,
                [
                    [
                        "X", dt.datetime(2025, 1, 1), "Main Item", "Item C",
                        30, None, None, None,
                    ],
                ],
            )
            n2 = upsert(db, load_rows(xlsx2), location_id="default", source="run1")
            self.assertEqual(n2, 1)

            con = sqlite3.connect(str(db))
            try:
                run1_items = sorted(
                    r[0] for r in con.execute(
                        "SELECT item FROM beo_prep_history WHERE source = 'run1'"
                    )
                )
                other = sorted(
                    r[0] for r in con.execute(
                        "SELECT item FROM beo_prep_history WHERE source = 'other_source'"
                    )
                )
            finally:
                con.close()

            self.assertEqual(run1_items, ["Item C"])
            self.assertEqual(other, ["Untouched"])

    def test_upsert_rolls_back_on_error(self) -> None:
        # Inject a row with item=None to force the NOT NULL constraint to bite,
        # then assert the existing rows for that source aren't lost.
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            db = tmpdir / "lariat.db"
            _make_db(db)
            xlsx = tmpdir / "fixture.xlsx"
            _make_fixture_xlsx(
                xlsx,
                [
                    [
                        "X", dt.datetime(2025, 1, 1), "Main Item", "Original",
                        10, None, None, None,
                    ],
                ],
            )
            upsert(db, load_rows(xlsx), location_id="default", source="r")

            bad_rows = [
                {
                    "client": None, "event_date": None, "type": None,
                    "item": None,  # violates NOT NULL after the DELETE
                    "amount_qty": None, "prep_day": None,
                    "pre_prep_notes": None, "plating_notes": None,
                },
            ]
            with self.assertRaises(sqlite3.IntegrityError):
                upsert(db, bad_rows, location_id="default", source="r")

            con = sqlite3.connect(str(db))
            try:
                items = sorted(
                    r[0] for r in con.execute(
                        "SELECT item FROM beo_prep_history WHERE source = 'r'"
                    )
                )
            finally:
                con.close()
            # Original row preserved by the rollback.
            self.assertEqual(items, ["Original"])


if __name__ == "__main__":
    unittest.main()
